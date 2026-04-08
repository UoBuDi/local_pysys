from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, Depends, Query, UploadFile, File, Form
from pydantic import BaseModel
import configparser
import logging
import pymysql
import hashlib
import os
import re
import uvicorn
import asyncio
import json
import requests
import base64
from io import BytesIO
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from config import load_config, save_config, get_database_config
from database import test_db_connection, create_db_connection, get_database_tables, table_exists, get_available_databases, create_db_pool, close_all_pools
from sync_service import generate_month_options, run_sync, stop_sync, sync_status
from split_match_service import SplitMatchService
from statistics_service import get_dashboard_statistics, run_statistics_task, update_task_status, get_latest_month_data, start_task_execution, end_task_execution, get_task_execution_history, get_check_data_connection
from cloud_portal_data_service import run_cloud_portal_data_sync, get_cloud_portal_data
from models import *

PASSWORD_SECRET_KEY = "cloud_portal_secret_key_2024"

def encrypt_password(password: str) -> str:
    encrypted = []
    for i, char in enumerate(password):
        key_char = PASSWORD_SECRET_KEY[i % len(PASSWORD_SECRET_KEY)]
        encrypted_char = chr(ord(char) ^ ord(key_char))
        encrypted.append(encrypted_char)
    return base64.b64encode(''.join(encrypted).encode()).decode()

def decrypt_password(encrypted_password: str) -> str:
    try:
        decoded = base64.b64decode(encrypted_password.encode()).decode()
        decrypted = []
        for i, char in enumerate(decoded):
            key_char = PASSWORD_SECRET_KEY[i % len(PASSWORD_SECRET_KEY)]
            decrypted_char = chr(ord(char) ^ ord(key_char))
            decrypted.append(decrypted_char)
        return ''.join(decrypted)
    except Exception as e:
        logger.debug(f"密码解密失败: {e}")
        return ''

# 从models导入所有数据模型
from models import (
    DepartmentCreateRequest, DepartmentUpdateRequest,
    UserCreateRequest, UserUpdateRequest,
    RoleCreateRequest, RoleUpdateRequest,
    MenuCreateRequest, MenuUpdateRequest,
    AssignRoleRequest, AssignMenuRequest
)

# 配置日志
log_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# 创建日志记录器
logger = logging.getLogger()
logger.setLevel(logging.WARNING)

# 清除默认处理器
logger.handlers.clear()

# 控制台处理器
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
logger.addHandler(console_handler)

# 文件处理器 - 普通日志
log_dir = os.path.join(os.path.dirname(__file__), 'logs')
os.makedirs(log_dir, exist_ok=True)
file_handler = logging.FileHandler(os.path.join(log_dir, 'app.log'), encoding='utf-8')
file_handler.setFormatter(log_formatter)
logger.addHandler(file_handler)

# 文件处理器 - 错误日志
error_handler = logging.FileHandler(os.path.join(log_dir, 'error.log'), encoding='utf-8')
error_handler.setLevel(logging.ERROR)
error_handler.setFormatter(log_formatter)
logger.addHandler(error_handler)

logger = logging.getLogger(__name__)

def format_time_value(time_val):
    """格式化时间值，处理datetime.time和timedelta类型"""
    if not time_val:
        return ''
    if hasattr(time_val, 'strftime'):
        return time_val.strftime('%H:%M:%S')
    if hasattr(time_val, 'total_seconds'):
        total_seconds = int(time_val.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return str(time_val)

app = FastAPI()

# 添加静态文件服务
static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# 添加CORS中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 数据模型
class DatabaseConfig(BaseModel):
    host: str
    port: int = 3306
    user: str
    password: str
    database: str
    charset: str = "utf8mb4"

class TestConnectionRequest(BaseModel):
    host: str
    port: int
    user: str
    password: str
    database: str
    charset: str

class StartSyncRequest(BaseModel):
    months: List[str]

class SyncStatusResponse(BaseModel):
    is_running: bool
    progress: float = 0.0
    message: str = ""

class LoginCredentials(BaseModel):
    username: str
    password: str

class RegisterCredentials(BaseModel):
    username: str
    password: str
    email: Optional[str] = None

class RoleCreateRequest(BaseModel):
    name: str
    code: str
    description: Optional[str] = None
    status: int = 1
    menu: Optional[List[dict]] = None

class RoleUpdateRequest(BaseModel):
    name: str
    code: str
    description: Optional[str] = None
    status: int = 1
    menu: Optional[List[dict]] = None

class AssignRoleRequest(BaseModel):
    user_id: int
    role_ids: List[int]

class AssignMenuRequest(BaseModel):
    role_id: int
    menu_ids: List[int]

class AssignPermissionRequest(BaseModel):
    role_id: int
    permission_ids: List[int]

class PermissionCreateRequest(BaseModel):
    code: str
    name: str
    module: str
    resource: str
    operation: str
    description: Optional[str] = None
    status: int = 1

class PermissionUpdateRequest(BaseModel):
    name: str
    module: str
    resource: str
    operation: str
    description: Optional[str] = None
    status: int = 1

# 加载配置
config = load_config()

# 从sync_service导入同步状态，避免重复定义
from sync_service import sync_status

# 初始化同步任务变量
sync_task = None

# 存储WebSocket连接
active_connections = []

# JWT令牌生成和验证相关代码
import jwt
from datetime import datetime, timedelta
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

# JWT配置
SECRET_KEY = os.environ.get("SECRET_KEY", "change-me")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_DAYS = 7

security = HTTPBearer()

def create_access_token(data: dict):
    """创建访问令牌"""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire, "type": "access"})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict):
    """创建刷新令牌"""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def decode_token(token: str):
    """解码令牌"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.PyJWTError:
        return None

def get_token_expires_at():
    """获取访问令牌过期时间戳（毫秒）"""
    return int((datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)).timestamp() * 1000)

def get_refresh_token_expires_at():
    """获取刷新令牌过期时间戳（毫秒）"""
    return int((datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)).timestamp() * 1000)

# 获取当前用户信息的依赖
def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """获取当前用户信息"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="无效的访问令牌")
        
        token_type = payload.get("type", "access")
        if token_type != "access":
            raise HTTPException(status_code=401, detail="请使用访问令牌")
        
        # 查询用户详细信息和权限
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到用户数据库")
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 获取用户信息
                cursor.execute("SELECT id, username, nickname, email, department_id, status FROM users WHERE username = %s", (username,))
                user = cursor.fetchone()
                
                if not user:
                    raise HTTPException(status_code=401, detail="用户不存在")
                
                # 获取用户角色
                cursor.execute("SELECT r.id, r.name, r.code FROM roles r JOIN user_roles ur ON r.id = ur.role_id WHERE ur.user_id = %s", (user['id'],))
                roles = cursor.fetchall()
                
                # 获取用户权限
                cursor.execute("""
                    SELECT DISTINCT m.permission 
                    FROM menus m 
                    JOIN role_menus rm ON m.id = rm.menu_id 
                    JOIN user_roles ur ON rm.role_id = ur.role_id 
                    WHERE ur.user_id = %s AND m.permission IS NOT NULL AND m.permission != ''
                """, (user['id'],))
                permissions = [row['permission'] for row in cursor.fetchall()]
                
                # 构造用户信息
                user_info = {
                    "id": user['id'],
                    "username": user['username'],
                    "nickname": user['nickname'],
                    "email": user['email'],
                    "department_id": user['department_id'],
                    "status": user['status'],
                    "roles": roles,
                    "permissions": permissions
                }
                
                return user_info
        finally:
            conn.close()
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="无效的访问令牌")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取用户信息失败: {e}")
        raise HTTPException(status_code=500, detail="获取用户信息失败")

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """验证令牌"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="无效的访问令牌")

def require_permission(permission: str):
    """权限验证依赖"""
    async def check_permission(user: dict = Depends(get_current_user)):
        if 'super_admin' in [r.get('code') for r in user.get('roles', [])]:
            return user
        
        if permission not in user.get('permissions', []):
            raise HTTPException(status_code=403, detail="没有权限执行此操作")
        return user
    return check_permission

@app.get("/api/config/")
async def get_config():
    """获取当前配置"""
    result = {}
    for section in config.sections():
        result[section] = {}
        for key, value in config.items(section):
            result[section][key] = value
    return result

@app.post("/api/config/")
async def save_config_endpoint(config_data: dict):
    """保存配置"""
    global config
    # 更新配置对象
    for section, section_data in config_data.items():
        if not config.has_section(section):
            config.add_section(section)
        for key, value in section_data.items():
            config.set(section, key, str(value))
    
    # 保存到文件
    if save_config(config):
        # 广播配置变更通知
        config_message = json.dumps({
            "type": "config_update",
            "timestamp": datetime.now().isoformat(),
            "message": "配置已更新",
            "config": {
                section: dict(config[section]) for section in config.sections()
            }
        })
        await manager.broadcast(config_message)
        return {"message": "配置保存成功"}
    else:
        return JSONResponse(status_code=500, content={"message": "配置保存失败"})

@app.post("/api/config/test-remote/")
async def test_remote_db_connection(request: TestConnectionRequest):
    """测试远程数据库连接"""
    success = test_db_connection(
        request.host, request.port, request.user,
        request.password, request.database, request.charset
    )
    if success:
        return {"message": "远程数据库连接成功"}
    else:
        return JSONResponse(status_code=400, content={"message": "远程数据库连接失败"})

@app.post("/api/config/test-local/")
async def test_local_db_connection(request: TestConnectionRequest):
    """测试本地数据库连接"""
    success = test_db_connection(
        request.host, request.port, request.user,
        request.password, request.database, request.charset
    )
    if success:
        return {"message": "本地数据库连接成功"}
    else:
        return JSONResponse(status_code=400, content={"message": "本地数据库连接失败"})

@app.get("/api/database/tables/remote/")
async def get_remote_database_tables():
    """获取远程数据库表列表"""
    connection = create_db_connection("REMOTE_DB", config)
    if not connection:
        return JSONResponse(status_code=500, content={"message": "无法连接到远程数据库"})
    
    try:
        tables = get_database_tables(connection)
        return {"tables": tables}
    finally:
        connection.close()

@app.get("/api/database/tables/local/")
async def get_local_database_tables():
    """获取本地数据库表列表"""
    connection = create_db_connection("LOCAL_DB", config)
    if not connection:
        return JSONResponse(status_code=500, content={"message": "无法连接到本地数据库"})
    
    try:
        tables = get_database_tables(connection)
        return {"tables": tables}
    finally:
        connection.close()

@app.get("/api/sync/months/")
async def get_sync_months():
    """获取可选月份列表"""
    months = generate_month_options()
    return {"code": 200, "message": "success", "data": months}

@app.post("/api/start-sync/")
async def start_sync(request: StartSyncRequest):
    """开始同步"""
    global sync_task
    if sync_status["is_running"]:
        return {"code": 400, "message": "同步已在进行中", "data": None}
    
    # 启动同步任务
    logger.info(f"开始同步任务，月份: {request.months}")
    sync_task = asyncio.create_task(run_sync(request.months, config))
    return {"code": 200, "message": "同步已启动", "data": None}

@app.post("/api/stop-sync/")
async def stop_sync_endpoint():
    """停止同步"""
    logger.info("收到停止同步指令")
    stop_sync()
    return {"code": 200, "message": "停止同步指令已发送", "data": None}

@app.post("/api/pause-sync/")
async def pause_sync_endpoint():
    """暂停同步"""
    logger.info("收到暂停同步指令")
    # 暂停同步逻辑（如果有的话）
    return {"code": 200, "message": "暂停同步指令已发送", "data": None}

@app.post("/api/force-stop-sync/")
async def force_stop_sync_endpoint():
    """强制停止同步"""
    logger.info("收到强制停止同步指令")
    stop_sync()  # 复用现有的停止同步逻辑
    return {"code": 200, "message": "强制停止同步指令已发送", "data": None}

@app.get("/api/sync/status/")
async def get_sync_status():
    """获取同步状态"""
    # 构建与前端期望匹配的状态数据格式
    status_data = {
        "status": "RUNNING" if sync_status["is_running"] else "IDLE",
        "progress": f"{sync_status['progress']:.1f}%" if isinstance(sync_status['progress'], (int, float)) else sync_status['progress'],
        "current_month": sync_status["current_month"] or "",
        "last_update": ""  # 可以根据需要添加实际的最后更新时间
    }
    return {"code": 200, "message": "success", "data": status_data}

# 添加分析页面数据接口
@app.get("/api/analysis/total")
async def get_analysis_total():
    """获取分析总览数据 - 从统计数据表获取"""
    try:
        stats = get_dashboard_statistics()
        
        if stats:
            return {
                "code": 200,
                "message": "success",
                "data": [
                    {
                        "type": "transactions",
                        "value": stats.get('total_transactions', 0)
                    },
                    {
                        "type": "amount",
                        "value": float(stats.get('total_amount', 0))
                    },
                    {
                        "type": "splitAmount",
                        "value": float(stats.get('total_split_amount', 0))
                    },
                    {
                        "type": "stations",
                        "value": stats.get('station_count', 0)
                    }
                ]
            }
        else:
            return {
                "code": 200,
                "message": "success",
                "data": [
                    {"type": "transactions", "value": 0},
                    {"type": "amount", "value": 0},
                    {"type": "splitAmount", "value": 0},
                    {"type": "stations", "value": 0}
                ]
            }
    except Exception as e:
        logger.error(f"获取分析总览数据失败: {e}")
        return {
            "code": 500,
            "message": f"获取数据失败: {str(e)}",
            "data": []
        }

@app.get("/api/analysis/userAccessSource")
async def get_user_access_source():
    """获取车型分布数据 - 从统计数据表获取"""
    try:
        stats = get_dashboard_statistics()
        
        if stats and stats.get('vehicle_types'):
            vehicle_types = stats['vehicle_types']
            data = []
            for item in vehicle_types:
                data.append({
                    "name": item.get('type', '未知'),
                    "value": item.get('count', 0)
                })
            return {"code": 200, "message": "success", "data": data}
        else:
            return {"code": 200, "message": "success", "data": []}
    except Exception as e:
        logger.error(f"获取车型分布数据失败: {e}")
        return {"code": 500, "message": f"获取数据失败: {str(e)}", "data": []}

@app.get("/api/analysis/weeklyUserActivity")
async def get_weekly_user_activity():
    """获取通行介质类型统计数据 - 从统计数据表获取"""
    try:
        stats = get_dashboard_statistics()
        
        if stats and stats.get('media_types'):
            media_data = stats['media_types']
            data = []
            for item in media_data:
                media_type = item.get('type', '未知')
                data.append({
                    "name": f"通行介质{media_type}",
                    "value": item.get('count', 0)
                })
            return {"code": 200, "message": "success", "data": data}
        else:
            return {"code": 200, "message": "success", "data": []}
    except Exception as e:
        logger.error(f"获取通行介质类型数据失败: {e}")
        return {"code": 500, "message": f"获取数据失败: {str(e)}", "data": []}

@app.get("/api/analysis/monthlySales")
async def get_monthly_sales():
    """获取省份数据统计数据 - 从统计数据表获取"""
    try:
        stats = get_dashboard_statistics()
        
        if stats and stats.get('province_data'):
            province_data = stats['province_data']
            data = []
            for item in province_data:
                province_name = item.get('province', '未知')
                count = item.get('count', 0)
                data.append({
                    "name": province_name,
                    "estimate": count,
                    "actual": count
                })
            
            return {"code": 200, "message": "success", "data": data}
        else:
            return {"code": 200, "message": "success", "data": []}
    except Exception as e:
        logger.error(f"获取省份数据统计数据失败: {e}")
        return {"code": 500, "message": f"获取数据失败: {str(e)}", "data": []}

@app.post("/api/analytics/route")
async def log_route_analytics(request_data: dict):
    """记录路由导航分析数据"""
    try:
        timestamp = request_data.get('timestamp')
        from_path = request_data.get('from')
        to_path = request_data.get('to')
        duration = request_data.get('duration')
        user_agent = request_data.get('userAgent')
        
        logger.info(f"路由导航: {from_path} -> {to_path}, 耗时: {duration}ms")
        
        return {"code": 200, "message": "success"}
    except Exception as e:
        logger.error(f"记录路由分析数据失败: {e}")
        return {"code": 500, "message": "记录失败"}

@app.get("/api/user/loginOut")
async def login_out():
    """用户登出"""
    return {
        "code": 200,
        "message": "success"
    }

class RefreshTokenRequest(BaseModel):
    refresh_token: str

@app.post("/api/token/refresh")
async def refresh_token(request: RefreshTokenRequest):
    """使用Refresh Token获取新的Access Token"""
    try:
        payload = decode_token(request.refresh_token)
        
        if not payload:
            return {"code": 401, "message": "刷新令牌已过期或无效"}
        
        token_type = payload.get("type")
        if token_type != "refresh":
            return {"code": 401, "message": "无效的刷新令牌类型"}
        
        username = payload.get("sub")
        if not username:
            return {"code": 401, "message": "无效的刷新令牌"}
        
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                cursor.execute("SELECT id, username, status FROM users WHERE username = %s", (username,))
                user = cursor.fetchone()
                
                if not user:
                    return {"code": 401, "message": "用户不存在"}
                
                if user['status'] != 1:
                    return {"code": 401, "message": "用户已被禁用"}
                
                new_access_token = create_access_token(data={"sub": username})
                new_refresh_token = create_refresh_token(data={"sub": username})
                
                return {
                    "code": 200,
                    "message": "刷新成功",
                    "data": {
                        "token": new_access_token,
                        "refreshToken": new_refresh_token,
                        "tokenType": "bearer",
                        "expiresAt": get_token_expires_at(),
                        "refreshExpiresAt": get_refresh_token_expires_at()
                    }
                }
        finally:
            conn.close()
            
    except Exception as e:
        logger.error(f"刷新令牌失败: {e}")
        return {"code": 500, "message": "刷新令牌失败"}

@app.post("/api/token/validate")
async def validate_token(request: RefreshTokenRequest):
    """验证Token是否有效"""
    try:
        payload = decode_token(request.refresh_token)
        
        if not payload:
            return {"code": 401, "message": "令牌已过期或无效", "data": {"valid": False}}
        
        username = payload.get("sub")
        if not username:
            return {"code": 401, "message": "无效的令牌", "data": {"valid": False}}
        
        return {
            "code": 200,
            "message": "令牌有效",
            "data": {
                "valid": True,
                "username": username,
                "type": payload.get("type", "access"),
                "exp": payload.get("exp")
            }
        }
    except Exception as e:
        logger.error(f"验证令牌失败: {e}")
        return {"code": 500, "message": "验证令牌失败", "data": {"valid": False}}

@app.post("/api/login/")
async def login(credentials: LoginCredentials):
    """用户登录验证"""
    try:
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到用户数据库"})
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 查询用户信息
                sql = "SELECT id, username, password FROM users WHERE username = %s"
                cursor.execute(sql, (credentials.username,))
                user = cursor.fetchone()
                
                if not user:
                    return {"code": 401, "message": "用户名或密码错误"}
                
                # 验证密码（MD5）
                provided = credentials.password.encode('utf-8')
                stored = user['password']
                md5_hash = hashlib.md5(provided).hexdigest()
                if md5_hash != stored:
                    return {"code": 401, "message": "用户名或密码错误"}
                
                # 查询用户角色
                cursor.execute("""
                    SELECT r.id, r.name, r.code 
                    FROM roles r 
                    JOIN user_roles ur ON r.id = ur.role_id 
                    WHERE ur.user_id = %s AND r.status = 1
                """, (user['id'],))
                roles = cursor.fetchall()
                
                # 构建角色列表
                role_list = [role['name'] for role in roles] if roles else []
                
                # 生成JWT令牌
                access_token = create_access_token(data={"sub": user['username']})
                refresh_token = create_refresh_token(data={"sub": user['username']})
                
                # 登录成功
                return {
                    "code": 200,
                    "message": "登录成功",
                    "data": {
                        "user": {
                            "id": user['id'],
                            "username": user['username'],
                            "roles": roles if roles else [],
                            "roleList": role_list
                        },
                        "token": access_token,
                        "refreshToken": refresh_token,
                        "tokenType": "bearer",
                        "expiresAt": get_token_expires_at(),
                        "refreshExpiresAt": get_refresh_token_expires_at()
                    }
                }
        finally:
            conn.close()
            
    except Exception as e:
        logger.error(f"登录验证失败：{e}")
        return {"code": 500, "message": "登录验证失败"}

@app.post("/api/register/")
async def register(credentials: RegisterCredentials):
    """用户注册"""
    logger.info(f"开始处理注册请求: username={credentials.username}")
    
    # 验证输入参数
    if not credentials.username:
        logger.warning("用户名为空")
        return {"code": 400, "message": "用户名不能为空"}
    
    if not credentials.password:
        logger.warning("密码为空")
        return {"code": 400, "message": "密码不能为空"}
    
    if len(credentials.password) < 6:
        logger.warning(f"密码长度不足: {len(credentials.password)}")
        return {"code": 400, "message": "密码长度不能少于6位"}
    
    connection = None
    try:
        connection = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not connection:
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到用户数据库"})
        logger.info("数据库连接成功")
        
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查用户名是否已存在
            check_sql = "SELECT id, username FROM users WHERE username = %s"
            logger.info(f"执行查询: {check_sql} with params ({credentials.username},)")
            cursor.execute(check_sql, (credentials.username,))
            existing_user = cursor.fetchone()
            
            if existing_user:
                logger.warning(f"用户名已存在: {credentials.username}")
                return {"code": 400, "message": "用户名已存在"}
            
            # 密码加密（MD5）
            logger.info("开始密码加密")
            hashed_password = hashlib.md5(credentials.password.encode('utf-8')).hexdigest()
            logger.info("密码加密完成")
            
            # 插入新用户
            insert_sql = """
            INSERT INTO users (username, password, email, status, created_at) 
            VALUES (%s, %s, %s, %s, NOW())
            """
            params = (
                credentials.username,
                hashed_password,
                credentials.email or None,
                1  # 默认启用状态
            )
            
            logger.info(f"执行插入: {insert_sql} with params {params}")
            cursor.execute(insert_sql, params)
            logger.info("插入操作完成")
            
            # 提交事务
            logger.info("开始提交事务")
            connection.commit()
            logger.info("事务提交完成")
            
            # 获取新插入用户的ID
            new_user_id = cursor.lastrowid
            logger.info(f"新用户ID: {new_user_id}")
            
            logger.info(f"用户注册成功: {credentials.username} (ID: {new_user_id})")
            
            return {
                "code": 200,
                "message": "注册成功",
                "data": {
                    "user": {
                        "id": new_user_id,
                        "username": credentials.username
                    }
                }
            }
            
    except pymysql.err.IntegrityError as e:
        if connection:
            connection.rollback()
        logger.error(f"数据库完整性错误: {e}", exc_info=True)
        return {"code": 500, "message": "注册失败，数据完整性错误"}
        
    except pymysql.err.OperationalError as e:
        logger.error(f"数据库连接错误: {e}", exc_info=True)
        return {"code": 500, "message": "注册失败，无法连接到数据库"}
        
    except Exception as e:
        if connection and connection.open:
            connection.rollback()
        logger.error(f"用户注册异常: {e}", exc_info=True)
        return {"code": 500, "message": "注册失败，请稍后重试"}
        
    finally:
        if connection and connection.open:
            connection.close()
            logger.info("数据库连接已关闭")

# WebSocket连接管理
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        try:
            self.active_connections.remove(websocket)
        except ValueError:
            pass

    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)

    async def broadcast(self, message: str):
        disconnected = []
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except Exception as e:
                logger.debug(f"广播消息失败，移除断开的连接: {e}")
                disconnected.append(connection)
        
        # 清理断开的连接
        for conn in disconnected:
            self.disconnect(conn)

manager = ConnectionManager()

class StatusConnectionManager:
    def __init__(self):
        self.frontend_connections: List[dict] = []
        self.gui_connections: List[dict] = []
        self._lock = asyncio.Lock()
    
    async def connect_frontend(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        async with self._lock:
            self.frontend_connections.append({
                "websocket": websocket,
                "client_id": client_id,
                "last_heartbeat": datetime.now()
            })
    
    async def connect_gui(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        async with self._lock:
            self.gui_connections.append({
                "websocket": websocket,
                "client_id": client_id,
                "last_heartbeat": datetime.now()
            })
    
    async def disconnect_frontend(self, websocket: WebSocket):
        async with self._lock:
            self.frontend_connections = [
                conn for conn in self.frontend_connections 
                if conn["websocket"] != websocket
            ]
    
    async def disconnect_gui(self, websocket: WebSocket):
        async with self._lock:
            self.gui_connections = [
                conn for conn in self.gui_connections 
                if conn["websocket"] != websocket
            ]
    
    async def update_heartbeat(self, websocket: WebSocket, client_type: str):
        async with self._lock:
            if client_type == "frontend":
                for conn in self.frontend_connections:
                    if conn["websocket"] == websocket:
                        conn["last_heartbeat"] = datetime.now()
                        break
            elif client_type == "gui":
                for conn in self.gui_connections:
                    if conn["websocket"] == websocket:
                        conn["last_heartbeat"] = datetime.now()
                        break
    
    async def broadcast_to_frontend(self, message: dict):
        message_str = json.dumps(message)
        async with self._lock:
            for conn in self.frontend_connections:
                try:
                    await conn["websocket"].send_text(message_str)
                except Exception as e:
                    logger.debug(f"WebSocket发送到前端失败: {e}")
                    continue
    
    async def broadcast_to_gui(self, message: dict):
        message_str = json.dumps(message)
        async with self._lock:
            for conn in self.gui_connections:
                try:
                    await conn["websocket"].send_text(message_str)
                except Exception as e:
                    logger.debug(f"WebSocket发送到GUI失败: {e}")
                    continue
    
    async def broadcast_all(self, message: dict):
        await self.broadcast_to_frontend(message)
        await self.broadcast_to_gui(message)
    
    def get_status(self):
        return {
            "frontend_count": len(self.frontend_connections),
            "gui_count": len(self.gui_connections),
            "frontend_clients": [
                {
                    "client_id": conn["client_id"],
                    "last_heartbeat": conn["last_heartbeat"].isoformat()
                }
                for conn in self.frontend_connections
            ],
            "gui_clients": [
                {
                    "client_id": conn["client_id"],
                    "last_heartbeat": conn["last_heartbeat"].isoformat()
                }
                for conn in self.gui_connections
            ]
        }

status_manager = StatusConnectionManager()

async def broadcast_status_periodically():
    while True:
        try:
            status = status_manager.get_status()
            await status_manager.broadcast_to_frontend({
                "type": "status_update",
                "data": status,
                "timestamp": datetime.now().isoformat()
            })
            await status_manager.broadcast_to_gui({
                "type": "status_update",
                "data": status,
                "timestamp": datetime.now().isoformat()
            })
        except Exception as e:
            logger.error(f"广播状态失败: {e}")
        await asyncio.sleep(5)

@app.websocket("/ws/status/{client_type}/{client_id}")
async def websocket_status(websocket: WebSocket, client_type: str, client_id: str):
    if client_type not in ["frontend", "gui"]:
        await websocket.close(code=4000, reason="Invalid client type")
        return
    
    if client_type == "frontend":
        await status_manager.connect_frontend(websocket, client_id)
    else:
        await status_manager.connect_gui(websocket, client_id)
    
    try:
        await websocket.send_text(json.dumps({
            "type": "connected",
            "client_type": client_type,
            "client_id": client_id,
            "message": f"{client_type}客户端已连接",
            "timestamp": datetime.now().isoformat()
        }))
        
        status = status_manager.get_status()
        await status_manager.broadcast_all({
            "type": "client_joined",
            "client_type": client_type,
            "client_id": client_id,
            "status": status,
            "timestamp": datetime.now().isoformat()
        })
        
        while True:
            data = await websocket.receive_text()
            try:
                message = json.loads(data)
                
                if message.get("type") == "heartbeat":
                    await status_manager.update_heartbeat(websocket, client_type)
                    await websocket.send_text(json.dumps({
                        "type": "heartbeat_ack",
                        "timestamp": datetime.now().isoformat()
                    }))
                elif message.get("type") == "ping":
                    await websocket.send_text(json.dumps({
                        "type": "pong",
                        "timestamp": datetime.now().isoformat()
                    }))
                elif message.get("type") == "get_status":
                    status = status_manager.get_status()
                    await websocket.send_text(json.dumps({
                        "type": "status_response",
                        "data": status,
                        "timestamp": datetime.now().isoformat()
                    }))
                else:
                    await status_manager.broadcast_all({
                        "type": "message",
                        "from_type": client_type,
                        "from_id": client_id,
                        "data": message,
                        "timestamp": datetime.now().isoformat()
                    })
            except json.JSONDecodeError:
                await websocket.send_text(json.dumps({
                    "type": "error",
                    "message": "Invalid JSON format"
                }))
    except WebSocketDisconnect:
        if client_type == "frontend":
            await status_manager.disconnect_frontend(websocket)
        else:
            await status_manager.disconnect_gui(websocket)
        
        status = status_manager.get_status()
        await status_manager.broadcast_all({
            "type": "client_left",
            "client_type": client_type,
            "client_id": client_id,
            "status": status,
            "timestamp": datetime.now().isoformat()
        })

@app.get("/api/ws/status")
async def get_ws_status():
    return {
        "code": 200,
        "message": "success",
        "data": status_manager.get_status()
    }

# WebSocket日志处理器
class WebSocketLogHandler(logging.Handler):
    def emit(self, record):
        try:
            log_entry = self.format(record)
            loop = asyncio.get_event_loop()
            if loop.is_running():
                loop.call_soon_threadsafe(
                    lambda: asyncio.ensure_future(
                        manager.broadcast(json.dumps({
                            "timestamp": datetime.now().isoformat(),
                            "level": record.levelname,
                            "message": log_entry
                        }))
                    )
                )
        except Exception as e:
            pass

# 注册WebSocket日志处理器
websocket_handler = WebSocketLogHandler()
websocket_handler.setLevel(logging.INFO)
websocket_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logging.getLogger().addHandler(websocket_handler)

@app.websocket("/ws/sync-progress/")
async def websocket_sync_progress(websocket: WebSocket):
    """WebSocket同步进度推送"""
    await manager.connect(websocket)
    try:
        while True:
            # 等待接收消息（这里我们主要是推送）
            data = await websocket.receive_text()
            # 可以在这里处理客户端发来的消息
    except WebSocketDisconnect:
        manager.disconnect(websocket)

@app.websocket("/ws/logs/")
async def websocket_logs(websocket: WebSocket):
    """WebSocket日志推送"""
    await manager.connect(websocket)
    try:
        # 发送初始日志
        await websocket.send_text(json.dumps({
            "timestamp": asyncio.get_event_loop().time(),
            "level": "INFO",
            "message": "已连接到日志服务"
        }))
        
        while True:
            # 等待接收消息
            data = await websocket.receive_text()
            # 可以在这里处理客户端发来的消息
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# 在同步过程中广播进度更新
async def broadcast_sync_progress():
    """定期广播同步进度"""
    while True:
        if sync_status["is_running"]:
            progress_message = json.dumps({
                "type": "progress",
                "progress": sync_status["progress"],
                "message": sync_status["message"]
            })
            await manager.broadcast(progress_message)
        await asyncio.sleep(1)  # 每秒更新一次

# 启动后台任务
@app.on_event("startup")
async def startup_event():
    asyncio.create_task(broadcast_sync_progress())
    asyncio.create_task(broadcast_status_periodically())

# 获取数据库连接的依赖项
def get_db():
    connection = create_db_connection("USER_DB", config)
    if not connection:
        raise HTTPException(status_code=500, detail="无法连接到用户数据库")
    try:
        yield connection
    finally:
        connection.close()

@app.get("/api/user/menus")
async def get_user_menus(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """获取用户菜单 - 根据用户的角色和菜单分配返回对应的菜单"""
    try:
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到数据库"})
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 获取当前用户信息
                token = credentials.credentials
                payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                username = payload.get("sub")
                
                if not username:
                    return {"code": 401, "message": "无效的访问令牌"}
                
                # 获取用户ID
                cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
                user = cursor.fetchone()
                
                if not user:
                    return {"code": 401, "message": "用户不存在"}
                
                user_id = user['id']
                
                # 获取用户的角色
                cursor.execute("""
                    SELECT DISTINCT r.id, r.name, r.code 
                    FROM roles r 
                    JOIN user_roles ur ON r.id = ur.role_id 
                    WHERE ur.user_id = %s AND r.status = 1
                """, (user_id,))
                roles = cursor.fetchall()
                
                if not roles:
                    # 如果用户没有角色，返回空菜单
                    return {"code": 200, "message": "success", "data": []}
                
                role_ids = [role['id'] for role in roles]
                
                # 获取用户可访问的菜单 - 使用安全的参数化查询
                placeholders = ','.join(['%s'] * len(role_ids))
                cursor.execute(f"""
                    SELECT DISTINCT m.* 
                    FROM menus m 
                    JOIN role_menus rm ON m.id = rm.menu_id 
                    WHERE rm.role_id IN ({placeholders}) AND m.status = 1 AND m.visible = 1
                    ORDER BY m.sort_order
                """, role_ids)
                menus = cursor.fetchall()
                
                # 转换字段名，确保前端兼容性
                converted_menus = []
                # 先建立菜单ID到菜单的映射，方便查找父菜单
                menu_map = {menu['id']: menu for menu in menus}
                
                for menu in menus:
                    # 跳过没有path的菜单（按钮权限）
                    if not menu.get('path'):
                        continue
                        
                    converted_menu = menu.copy()
                    # 同时保留 sort_order 和 sortOrder
                    if 'sort_order' in converted_menu:
                        converted_menu['sortOrder'] = converted_menu['sort_order']
                    # 转换 parent_id 为 parentId
                    if 'parent_id' in converted_menu:
                        converted_menu['parentId'] = converted_menu['parent_id']
                    
                    # 对于有子菜单的顶级菜单，设置component为#
                    if converted_menu.get('parent_id') == 0:
                        converted_menu['component'] = '#'
                    # 对于子菜单，根据路径自动设置component
                    elif converted_menu.get('parent_id') != 0:
                        full_path = converted_menu.get('path', '')
                        # 获取父菜单路径
                        parent_menu = menu_map.get(converted_menu.get('parent_id'))
                        if parent_menu:
                            parent_path = parent_menu.get('path', '')
                            # 将完整路径转换为相对路径
                            if full_path.startswith(parent_path + '/'):
                                converted_menu['path'] = full_path[len(parent_path) + 1:]
                            elif full_path.startswith(parent_path):
                                converted_menu['path'] = full_path[len(parent_path):].lstrip('/')
                        
                        # 如果数据库中已经有component字段，优先使用
                        if converted_menu.get('component'):
                            component = converted_menu['component']
                            if component == 'Layout':
                                converted_menu['component'] = '#'
                            elif component.startswith('/'):
                                converted_menu['component'] = component[1:]
                            # 修复 system-tools 为 SystemTools
                            elif component.startswith('system-tools/'):
                                converted_menu['component'] = 'SystemTools' + component[12:]
                        else:
                            # 根据路径映射到正确的组件
                            path = converted_menu.get('path', '')
                            if path == 'analysis':
                                converted_menu['component'] = 'Dashboard/Analysis'
                            elif path == 'workplace':
                                converted_menu['component'] = 'Dashboard/Workplace'
                            elif path == 'split-match':
                                converted_menu['component'] = 'SystemTools/SplitMatch'
                            elif path == 'detail-query':
                                converted_menu['component'] = 'SystemTools/DetailQuery'
                            elif path == 'path-match':
                                converted_menu['component'] = 'SystemTools/PathMatch'
                            elif path == 'config' or path == 'sync-config':
                                converted_menu['component'] = 'SystemTools/SyncConfig'
                            elif path == 'control' or path == 'sync-control':
                                converted_menu['component'] = 'SystemTools/SyncControl'
                            elif path == 'params-config':
                                converted_menu['component'] = 'SystemTools/ParamsConfig'
                            elif path == 'department':
                                converted_menu['component'] = 'Authorization/Department/Department'
                            elif path == 'user':
                                converted_menu['component'] = 'Authorization/User/User'
                            elif path == 'menu':
                                converted_menu['component'] = 'Authorization/Menu/Menu'
                            elif path == 'role':
                                converted_menu['component'] = 'Authorization/Role/Role'
                            elif path == 'scheduled-tasks':
                                converted_menu['component'] = 'SystemTools/ScheduledTasks/ScheduledTasks'
                            elif path == 'monitor-center':
                                converted_menu['component'] = 'ParentLayout'
                            elif path == 'special-records':
                                converted_menu['component'] = 'DataRecords/SpecialRecords'
                            elif path == 'event-records':
                                converted_menu['component'] = 'DataRecords/EventRecords'
                            elif path == 'scheduling':
                                converted_menu['component'] = 'ParentLayout'
                            elif path == 'calendar':
                                converted_menu['component'] = 'DataRecords/Scheduling/Calendar'
                            elif path == 'groups':
                                converted_menu['component'] = 'DataRecords/Scheduling/Groups'
                            elif path == 'staff':
                                converted_menu['component'] = 'DataRecords/Scheduling/Staff'
                            elif path == 'shifts':
                                converted_menu['component'] = 'DataRecords/Scheduling/Shifts'
                            elif path == 'manage-center':
                                converted_menu['component'] = 'DataRecords/ManageCenter'
                    # 添加 meta 字段，将中文菜单名映射为国际化key
                    name = converted_menu.get('name', '')
                    title_map = {
                        '数据查询': 'router.dataQuery',
                        '拆分匹配': 'router.splitMatch',
                        '详单查询': 'router.detailQuery',
                        '路径匹配': 'router.pathMatch',
                        '系统管理': 'router.authorization',
                        '用户管理': 'router.user',
                        '角色管理': 'router.role',
                        '菜单管理': 'router.menuManagement',
                        '部门管理': 'router.department',
                        '首页': 'router.dashboard',
                        '工作台': 'router.workplace',
                        '分析页': 'router.analysis',
                        '同步管理': 'router.syncManage',
                        '同步配置': 'router.syncConfig',
                        '同步控制': 'router.syncControl',
                        '系统工具': 'router.systemTools',
                        '定时任务': 'router.scheduledTasks',
                        '数据记录': 'router.dataRecords',
                        '监控中心': 'router.monitorCenter',
                        '特情记录': 'router.specialRecords',
                        '事件记录': 'router.eventRecords',
                        '排班功能': 'router.scheduling',
                        '排班管理': 'router.scheduling',
                        '排班日历': 'router.schedulingCalendar',
                        '班组管理': 'router.schedulingGroups',
                        '人员管理': 'router.schedulingStaff',
                        '班次管理': 'router.schedulingShifts',
                        '管理中心': 'router.manageCenter'
                    }
                    converted_menu['meta'] = {
                        'title': title_map.get(name, name),
                        'icon': converted_menu.get('icon', ''),
                        'hidden': not converted_menu.get('visible', 1),
                        'alwaysShow': bool(converted_menu.get('always_show', 0)),
                        'noCache': bool(converted_menu.get('no_cache', 0)),
                        'breadcrumb': True,
                        'affix': bool(converted_menu.get('affix', 0)),
                        'noTagsView': False,
                        'canTo': bool(converted_menu.get('can_to', 1)),
                        'activeMenu': ''
                    }
                    # 添加 name 字段
                    if not converted_menu.get('name'):
                        converted_menu['name'] = converted_menu.get('path', '').replace('/', '').replace('-', '')
                    converted_menus.append(converted_menu)
                
                # 构造树形结构
                menu_tree = build_menu_tree(converted_menus)
                return {"code": 200, "message": "success", "data": menu_tree}
                
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"获取用户菜单失败: {e}")
        import traceback
        traceback.print_exc()
        return {"code": 500, "message": "获取用户菜单失败"}

# 部门管理API
@app.get("/api/departments/")
async def get_departments(
    pageIndex: Optional[int] = None,
    pageSize: Optional[int] = None,
    db = Depends(get_db)
):
    """获取部门列表"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM departments ORDER BY sort_order")
            departments = cursor.fetchall()
            
            # 如果提供了分页参数，返回分页数据
            if pageIndex and pageSize:
                total = len(departments)
                start = (pageIndex - 1) * pageSize
                end = start + pageSize
                paginated_departments = departments[start:end]
                return {
                    "code": 200, 
                    "message": "success", 
                    "data": {
                        "list": paginated_departments,
                        "total": total,
                        "pageIndex": pageIndex,
                        "pageSize": pageSize
                    }
                }
            # 否则返回所有数据
            return {"code": 200, "message": "success", "data": departments}
    except Exception as e:
        logger.error(f"获取部门列表失败: {e}")
        return {"code": 500, "message": "获取部门列表失败"}

@app.post("/api/departments/")
async def create_department(department: DepartmentCreateRequest, db = Depends(get_db)):
    """创建部门"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查部门名称是否已存在
            cursor.execute("SELECT id FROM departments WHERE name = %s", (department.name,))
            if cursor.fetchone():
                return {"code": 400, "message": "部门名称已存在"}
            
            # 插入新部门
            cursor.execute(
                "INSERT INTO departments (name, parent_id, sort_order, status) VALUES (%s, %s, %s, %s)",
                (department.name, department.parent_id, department.sort_order, department.status)
            )
            db.commit()
            return {"code": 200, "message": "部门创建成功"}
    except Exception as e:
        logger.error(f"创建部门失败: {e}")
        return {"code": 500, "message": "创建部门失败"}

@app.put("/api/departments/{department_id}")
async def update_department(department_id: int, department: DepartmentUpdateRequest, db = Depends(get_db)):
    """更新部门"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查部门是否存在
            cursor.execute("SELECT id FROM departments WHERE id = %s", (department_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "部门不存在"}
            
            # 检查部门名称是否已存在（排除自己）
            cursor.execute("SELECT id FROM departments WHERE name = %s AND id != %s", (department.name, department_id))
            if cursor.fetchone():
                return {"code": 400, "message": "部门名称已存在"}
            
            # 更新部门信息
            cursor.execute(
                "UPDATE departments SET name=%s, parent_id=%s, sort_order=%s, status=%s WHERE id=%s",
                (department.name, department.parent_id, department.sort_order, department.status, department_id)
            )
            db.commit()
            return {"code": 200, "message": "部门更新成功"}
    except Exception as e:
        logger.error(f"更新部门失败: {e}")
        return {"code": 500, "message": "更新部门失败"}

@app.delete("/api/departments/{department_id}")
async def delete_department(department_id: int, db = Depends(get_db)):
    """删除部门"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查部门是否存在
            cursor.execute("SELECT id FROM departments WHERE id = %s", (department_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "部门不存在"}
            
            # 检查是否有子部门
            cursor.execute("SELECT id FROM departments WHERE parent_id = %s", (department_id,))
            if cursor.fetchone():
                return {"code": 400, "message": "该部门下有子部门，无法删除"}
            
            # 检查是否有用户属于该部门
            cursor.execute("SELECT id FROM users WHERE department_id = %s", (department_id,))
            if cursor.fetchone():
                return {"code": 400, "message": "该部门下有用户，无法删除"}
            
            # 删除部门
            cursor.execute("DELETE FROM departments WHERE id = %s", (department_id,))
            db.commit()
            return {"code": 200, "message": "部门删除成功"}
    except Exception as e:
        logger.error(f"删除部门失败: {e}")
        return {"code": 500, "message": "删除部门失败"}

# 用户管理API
@app.get("/api/users/")
async def get_users(
    department_id: Optional[str] = None,
    pageIndex: int = 1,
    pageSize: int = 10,
    db = Depends(get_db)
):
    """获取用户列表"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 处理department_id：空字符串转为None，否则转为整数
            dept_id = int(department_id) if department_id and department_id.isdigit() else None
            
            # 构建查询条件
            if dept_id:
                cursor.execute("""
                    SELECT u.*, d.name as department_name 
                    FROM users u 
                    LEFT JOIN departments d ON u.department_id = d.id 
                    WHERE u.department_id = %s 
                    ORDER BY u.id
                """, (dept_id,))
            else:
                cursor.execute("""
                    SELECT u.*, d.name as department_name 
                    FROM users u 
                    LEFT JOIN departments d ON u.department_id = d.id 
                    ORDER BY u.id
                """)
            
            # 获取所有数据
            all_users = cursor.fetchall()
            
            # 为每个用户获取角色信息
            for user in all_users:
                cursor.execute("SELECT role_id FROM user_roles WHERE user_id = %s", (user['id'],))
                user_roles = cursor.fetchall()
                user['roles'] = [role['role_id'] for role in user_roles]
            
            # 实现分页
            total = len(all_users)
            start = (pageIndex - 1) * pageSize
            end = start + pageSize
            paginated_users = all_users[start:end]
            
            return {
                "code": 200, 
                "message": "success", 
                "data": {
                    "list": paginated_users,
                    "total": total,
                    "pageIndex": pageIndex,
                    "pageSize": pageSize
                }
            }
    except Exception as e:
        logger.error(f"获取用户列表失败: {e}")
        return {"code": 500, "message": "获取用户列表失败"}

@app.post("/api/users/")
async def create_user(user: UserCreateRequest, db = Depends(get_db)):
    """创建用户"""
    try:
        logger.info(f"开始创建用户，接收数据: {user.dict()}")
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查用户名是否已存在
            cursor.execute("SELECT id FROM users WHERE username = %s", (user.username,))
            if cursor.fetchone():
                logger.warning(f"用户名已存在: {user.username}")
                return {"code": 400, "message": "用户名已存在"}
            
            # 密码加密
            hashed_password = hashlib.md5(user.password.encode('utf-8')).hexdigest()
            
            # 插入新用户
            logger.info(f"执行创建用户SQL，参数: username={user.username}, nickname={user.nickname}, email={user.email}, department_id={user.department_id}, status={user.status}")
            cursor.execute(
                "INSERT INTO users (username, password, nickname, email, department_id, status) VALUES (%s, %s, %s, %s, %s, %s)",
                (user.username, hashed_password, user.nickname, user.email, user.department_id, user.status)
            )
            db.commit()
            logger.info(f"用户创建成功，生成ID: {cursor.lastrowid}")
            return {"code": 200, "message": "用户创建成功"}
    except Exception as e:
        logger.error(f"创建用户失败: {e}", exc_info=True)
        return {"code": 500, "message": "创建用户失败"}

@app.put("/api/users/{user_id}")
async def update_user(user_id: int, user: UserUpdateRequest, db = Depends(get_db)):
    """更新用户"""
    try:
        logger.info(f"开始更新用户，用户ID: {user_id}, 接收数据: {user.dict()}")
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查用户是否存在
            cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
            if not cursor.fetchone():
                logger.warning(f"用户不存在: {user_id}")
                return {"code": 404, "message": "用户不存在"}
            
            # 检查用户名是否已存在（排除自己）
            cursor.execute("SELECT id FROM users WHERE username = %s AND id != %s", (user.username, user_id))
            if cursor.fetchone():
                logger.warning(f"用户名已存在: {user.username}")
                return {"code": 400, "message": "用户名已存在"}
            
            # 更新用户信息
            if user.password:
                # 密码加密（MD5）
                hashed_password = hashlib.md5(user.password.encode('utf-8')).hexdigest()
                logger.info(f"新密码已加密")
                
                logger.info(f"执行更新用户（带密码）SQL，参数: username={user.username}, nickname={user.nickname}, email={user.email}, department_id={user.department_id}, status={user.status}, id={user_id}")
                cursor.execute(
                    "UPDATE users SET username=%s, password=%s, nickname=%s, email=%s, department_id=%s, status=%s WHERE id=%s",
                    (user.username, hashed_password, user.nickname, user.email, user.department_id, user.status, user_id)
                )
            else:
                # 如果未提供新密码，则不更新密码字段
                logger.info(f"执行更新用户（不带密码）SQL，参数: username={user.username}, nickname={user.nickname}, email={user.email}, department_id={user.department_id}, status={user.status}, id={user_id}")
                cursor.execute(
                    "UPDATE users SET username=%s, nickname=%s, email=%s, department_id=%s, status=%s WHERE id=%s",
                    (user.username, user.nickname, user.email, user.department_id, user.status, user_id)
                )
            db.commit()
            logger.info(f"用户更新成功，用户ID: {user_id}, 影响行数: {cursor.rowcount}")
            return {"code": 200, "message": "用户更新成功"}
    except Exception as e:
        logger.error(f"更新用户失败: {e}", exc_info=True)
        return {"code": 500, "message": "更新用户失败"}

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, db = Depends(get_db)):
    """删除用户"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查用户是否存在
            cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "用户不存在"}
            
            # 删除用户角色关联
            cursor.execute("DELETE FROM user_roles WHERE user_id = %s", (user_id,))
            
            # 删除用户
            cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
            db.commit()
            return {"code": 200, "message": "用户删除成功"}
    except Exception as e:
        logger.error(f"删除用户失败: {e}")
        return {"code": 500, "message": "删除用户失败"}

# 角色管理API
@app.get("/api/roles/")
async def get_roles(db = Depends(get_db)):
    """获取角色列表"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM roles ORDER BY id")
            roles = cursor.fetchall()
            return {"code": 200, "message": "success", "data": {"list": roles, "total": len(roles)}}
    except Exception as e:
        logger.error(f"获取角色列表失败: {e}")
        return {"code": 500, "message": "获取角色列表失败"}

@app.post("/api/roles/")
async def create_role(role: RoleCreateRequest, db = Depends(get_db)):
    """创建角色"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查角色编码是否已存在
            cursor.execute("SELECT id FROM roles WHERE code = %s", (role.code,))
            if cursor.fetchone():
                return {"code": 400, "message": "角色编码已存在"}
            
            # 插入新角色
            cursor.execute(
                "INSERT INTO roles (name, code, description, status) VALUES (%s, %s, %s, %s)",
                (role.name, role.code, role.description, role.status)
            )
            role_id = cursor.lastrowid
            
            # 如果包含菜单数据，添加菜单分配
            if role.menu is not None:
                # 递归添加菜单分配
                def add_menu_recursively(menus):
                    for menu in menus:
                        cursor.execute(
                            "INSERT INTO role_menus (role_id, menu_id) VALUES (%s, %s)",
                            (role_id, menu['id'])
                        )
                        # 递归处理子菜单
                        if 'children' in menu and menu['children']:
                            add_menu_recursively(menu['children'])
                
                add_menu_recursively(role.menu)
            
            db.commit()
            return {"code": 200, "message": "角色创建成功"}
    except Exception as e:
        logger.error(f"创建角色失败: {e}")
        return {"code": 500, "message": "创建角色失败"}

@app.put("/api/roles/{role_id}")
async def update_role(role_id: int, role: RoleUpdateRequest, db = Depends(get_db)):
    """更新角色"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查角色是否存在
            cursor.execute("SELECT id FROM roles WHERE id = %s", (role_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "角色不存在"}
            
            # 检查角色编码是否已存在（排除自己）
            cursor.execute("SELECT id FROM roles WHERE code = %s AND id != %s", (role.code, role_id))
            if cursor.fetchone():
                return {"code": 400, "message": "角色编码已存在"}
            
            # 更新角色信息
            cursor.execute(
                "UPDATE roles SET name=%s, code=%s, description=%s, status=%s WHERE id=%s",
                (role.name, role.code, role.description, role.status, role_id)
            )
            
            # 如果包含菜单数据，更新菜单分配
            if role.menu is not None:
                # 先删除角色原有的菜单
                cursor.execute("DELETE FROM role_menus WHERE role_id = %s", (role_id,))
                
                # 递归添加菜单分配
                def add_menu_recursively(menus):
                    for menu in menus:
                        cursor.execute(
                            "INSERT INTO role_menus (role_id, menu_id) VALUES (%s, %s)",
                            (role_id, menu['id'])
                        )
                        # 递归处理子菜单
                        if 'children' in menu and menu['children']:
                            add_menu_recursively(menu['children'])
                
                add_menu_recursively(role.menu)
            
            db.commit()
            return {"code": 200, "message": "角色更新成功"}
    except Exception as e:
        logger.error(f"更新角色失败: {e}")
        return {"code": 500, "message": "更新角色失败"}

@app.delete("/api/roles/{role_id}")
async def delete_role(role_id: int, db = Depends(get_db)):
    """删除角色"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查角色是否存在
            cursor.execute("SELECT id FROM roles WHERE id = %s", (role_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "角色不存在"}
            
            # 删除角色菜单关联
            cursor.execute("DELETE FROM role_menus WHERE role_id = %s", (role_id,))
            
            # 删除用户角色关联
            cursor.execute("DELETE FROM user_roles WHERE role_id = %s", (role_id,))
            
            # 删除角色
            cursor.execute("DELETE FROM roles WHERE id = %s", (role_id,))
            db.commit()
            return {"code": 200, "message": "角色删除成功"}
    except Exception as e:
        logger.error(f"删除角色失败: {e}")
        return {"code": 500, "message": "删除角色失败"}

# 菜单管理API
@app.get("/api/menus/")
async def get_menus(db = Depends(get_db)):
    """获取菜单列表"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM menus ORDER BY sort_order, id")
            menus = cursor.fetchall()
            
            # 转换字段名，确保前端兼容性
            converted_menus = []
            for menu in menus:
                converted_menu = menu.copy()
                # 同时保留 sort_order 和 sortOrder
                if 'sort_order' in converted_menu:
                    converted_menu['sortOrder'] = converted_menu['sort_order']
                # 转换 parent_id 为 parentId
                if 'parent_id' in converted_menu:
                    converted_menu['parentId'] = converted_menu['parent_id']
                # 添加 meta 字段
                converted_menu['meta'] = {
                    'title': converted_menu.get('name', ''),
                    'icon': converted_menu.get('icon', ''),
                    'hidden': not converted_menu.get('visible', 1),
                    'alwaysShow': bool(converted_menu.get('always_show', 0)),
                    'noCache': bool(converted_menu.get('no_cache', 0)),
                    'breadcrumb': True,
                    'affix': bool(converted_menu.get('affix', 0)),
                    'noTagsView': False,
                    'canTo': bool(converted_menu.get('can_to', 1)),
                    'activeMenu': ''
                }
                # 添加 type 字段
                if 'type' in converted_menu:
                    converted_menu['type'] = converted_menu['type']
                converted_menus.append(converted_menu)
            
            # 返回扁平列表，菜单管理页面自己处理树形结构
            return {"code": 200, "message": "success", "data": {"list": converted_menus}}
    except Exception as e:
        logger.error(f"获取菜单列表失败: {e}")
        import traceback
        traceback.print_exc()
        return {"code": 500, "message": "获取菜单列表失败"}

def build_menu_tree(menus, parent_id=0):
    """构建菜单树形结构"""
    tree = []
    for menu in menus:
        if menu.get('parent_id', 0) == parent_id:
            children = build_menu_tree(menus, menu['id'])
            if children:
                menu['children'] = children
            tree.append(menu)
    tree.sort(key=lambda x: x.get('sort_order', 0))
    return tree

@app.post("/api/menus/")
async def create_menu(menu_data: dict, db = Depends(get_db)):
    """创建菜单"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 提取字段，支持前端发送的多种格式
            parent_id = menu_data.get('parentId', menu_data.get('parent_id', 0))
            name = menu_data.get('name', menu_data.get('meta', {}).get('title', ''))
            icon = menu_data.get('icon', menu_data.get('meta', {}).get('icon', ''))
            path = menu_data.get('path', '')
            component = menu_data.get('component', '')
            permission = menu_data.get('permission', '')
            sort_order = menu_data.get('sortOrder', menu_data.get('sort_order', 0))
            status = menu_data.get('status', 1)
            visible = menu_data.get('visible', 1)
            menu_type = menu_data.get('type', 1)
            always_show = 1 if menu_data.get('meta', {}).get('alwaysShow', False) else 0
            no_cache = 1 if menu_data.get('meta', {}).get('noCache', False) else 0
            affix = 1 if menu_data.get('meta', {}).get('affix', False) else 0
            can_to = 1 if menu_data.get('meta', {}).get('canTo', True) else 0
            
            # 插入新菜单
            cursor.execute(
                """INSERT INTO menus (parent_id, name, icon, path, component, permission, sort_order, status, visible, type, always_show, no_cache, affix, can_to) 
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (parent_id, name, icon, path, component, permission, sort_order, status, visible, menu_type, always_show, no_cache, affix, can_to)
            )
            db.commit()
            return {"code": 200, "message": "菜单创建成功"}
    except Exception as e:
        logger.error(f"创建菜单失败: {e}")
        import traceback
        traceback.print_exc()
        return {"code": 500, "message": f"创建菜单失败: {str(e)}"}

@app.put("/api/menus/{menu_id}")
async def update_menu(menu_id: int, menu_data: dict, db = Depends(get_db)):
    """更新菜单"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查菜单是否存在
            cursor.execute("SELECT id FROM menus WHERE id = %s", (menu_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "菜单不存在"}
            
            # 提取字段，支持前端发送的多种格式
            parent_id = menu_data.get('parentId', menu_data.get('parent_id', 0))
            name = menu_data.get('name', menu_data.get('meta', {}).get('title', ''))
            icon = menu_data.get('icon', menu_data.get('meta', {}).get('icon', ''))
            path = menu_data.get('path', '')
            component = menu_data.get('component', '')
            permission = menu_data.get('permission', '')
            sort_order = menu_data.get('sortOrder', menu_data.get('sort_order', 0))
            status = menu_data.get('status', 1)
            visible = menu_data.get('visible', 1)
            menu_type = menu_data.get('type', 1)
            always_show = 1 if menu_data.get('meta', {}).get('alwaysShow', False) else 0
            no_cache = 1 if menu_data.get('meta', {}).get('noCache', False) else 0
            affix = 1 if menu_data.get('meta', {}).get('affix', False) else 0
            can_to = 1 if menu_data.get('meta', {}).get('canTo', True) else 0
            
            # 更新菜单信息
            cursor.execute(
                """UPDATE menus SET parent_id=%s, name=%s, icon=%s, path=%s, component=%s, permission=%s, 
                   sort_order=%s, status=%s, visible=%s, type=%s, always_show=%s, no_cache=%s, affix=%s, can_to=%s WHERE id=%s""",
                (parent_id, name, icon, path, component, permission, sort_order, status, visible, menu_type, always_show, no_cache, affix, can_to, menu_id)
            )
            db.commit()
            return {"code": 200, "message": "菜单更新成功"}
    except Exception as e:
        logger.error(f"更新菜单失败: {e}")
        import traceback
        traceback.print_exc()
        return {"code": 500, "message": f"更新菜单失败: {str(e)}"}

@app.delete("/api/menus/{menu_id}")
async def delete_menu(menu_id: int, db = Depends(get_db)):
    """删除菜单"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 检查菜单是否存在
            cursor.execute("SELECT id FROM menus WHERE id = %s", (menu_id,))
            if not cursor.fetchone():
                return {"code": 404, "message": "菜单不存在"}
            
            # 检查是否有子菜单
            cursor.execute("SELECT id FROM menus WHERE parent_id = %s", (menu_id,))
            if cursor.fetchone():
                return {"code": 400, "message": "该菜单下有子菜单，无法删除"}
            
            # 删除角色菜单关联
            cursor.execute("DELETE FROM role_menus WHERE menu_id = %s", (menu_id,))
            
            # 删除菜单
            cursor.execute("DELETE FROM menus WHERE id = %s", (menu_id,))
            db.commit()
            return {"code": 200, "message": "菜单删除成功"}
    except Exception as e:
        logger.error(f"删除菜单失败: {e}")
        return {"code": 500, "message": "删除菜单失败"}

# 用户角色分配API
@app.post("/api/user-roles/")
async def assign_user_roles(assign_request: AssignRoleRequest, db = Depends(get_db)):
    """为用户分配角色"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 先删除用户原有的角色
            cursor.execute("DELETE FROM user_roles WHERE user_id = %s", (assign_request.user_id,))
            
            # 添加新的角色分配
            for role_id in assign_request.role_ids:
                cursor.execute(
                    "INSERT INTO user_roles (user_id, role_id) VALUES (%s, %s)",
                    (assign_request.user_id, role_id)
                )
            db.commit()
            return {"code": 200, "message": "角色分配成功"}
    except Exception as e:
        logger.error(f"角色分配失败: {e}")
        return {"code": 500, "message": "角色分配失败"}

# 获取用户已分配角色的API
@app.get("/api/users/{user_id}/roles")
async def get_user_roles(user_id: int, db = Depends(get_db)):
    """获取用户已分配的角色"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT role_id FROM user_roles WHERE user_id = %s", (user_id,))
            role_ids = [row['role_id'] for row in cursor.fetchall()]
            return {"code": 200, "message": "success", "data": role_ids}
    except Exception as e:
        logger.error(f"获取用户角色失败: {e}")
        return {"code": 500, "message": "获取用户角色失败"}

# 角色菜单分配API
@app.post("/api/role-menus/")
async def assign_role_menus(assign_request: AssignMenuRequest, db = Depends(get_db)):
    """为角色分配菜单"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            # 先删除角色原有的菜单
            cursor.execute("DELETE FROM role_menus WHERE role_id = %s", (assign_request.role_id,))
            
            # 添加新的菜单分配
            for menu_id in assign_request.menu_ids:
                cursor.execute(
                    "INSERT INTO role_menus (role_id, menu_id) VALUES (%s, %s)",
                    (assign_request.role_id, menu_id)
                )
            db.commit()
            return {"code": 200, "message": "菜单分配成功"}
    except Exception as e:
        logger.error(f"菜单分配失败: {e}")
        return {"code": 500, "message": "菜单分配失败"}

# 获取角色已分配菜单的API
@app.get("/api/roles/{role_id}/menus")
async def get_role_menus(role_id: int, db = Depends(get_db)):
    """获取角色已分配的菜单"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT menu_id FROM role_menus WHERE role_id = %s", (role_id,))
            menu_ids = [row['menu_id'] for row in cursor.fetchall()]
            return {"code": 200, "message": "success", "data": menu_ids}
    except Exception as e:
        logger.error(f"获取角色菜单失败: {e}")
        return {"code": 500, "message": "获取角色菜单失败"}

# 权限点管理API
@app.get("/api/permissions/")
async def get_permissions(db = Depends(get_db)):
    """获取权限点列表"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM permissions ORDER BY module, resource, operation")
            permissions = cursor.fetchall()
            return {"code": 200, "message": "success", "data": permissions}
    except Exception as e:
        logger.error(f"获取权限点失败: {e}")
        return {"code": 500, "message": "获取权限点失败"}

@app.post("/api/permissions/")
async def create_permission(permission: PermissionCreateRequest, db = Depends(get_db)):
    """创建权限点"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                "INSERT INTO permissions (code, name, module, resource, operation, description, status) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (permission.code, permission.name, permission.module, permission.resource, permission.operation, permission.description, permission.status)
            )
            db.commit()
            return {"code": 200, "message": "权限点创建成功"}
    except Exception as e:
        logger.error(f"创建权限点失败: {e}")
        return {"code": 500, "message": "创建权限点失败"}

@app.put("/api/permissions/{permission_id}")
async def update_permission(permission_id: int, permission: PermissionUpdateRequest, db = Depends(get_db)):
    """更新权限点"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                "UPDATE permissions SET name = %s, module = %s, resource = %s, operation = %s, description = %s, status = %s WHERE id = %s",
                (permission.name, permission.module, permission.resource, permission.operation, permission.description, permission.status, permission_id)
            )
            db.commit()
            return {"code": 200, "message": "权限点更新成功"}
    except Exception as e:
        logger.error(f"更新权限点失败: {e}")
        return {"code": 500, "message": "更新权限点失败"}

@app.delete("/api/permissions/{permission_id}")
async def delete_permission(permission_id: int, db = Depends(get_db)):
    """删除权限点"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("DELETE FROM permissions WHERE id = %s", (permission_id,))
            db.commit()
            return {"code": 200, "message": "权限点删除成功"}
    except Exception as e:
        logger.error(f"删除权限点失败: {e}")
        return {"code": 500, "message": "删除权限点失败"}

# 角色权限分配API
@app.get("/api/roles/{role_id}/permissions")
async def get_role_permissions(role_id: int, db = Depends(get_db)):
    """获取角色已分配的权限点"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT permission_id FROM role_permissions WHERE role_id = %s", (role_id,))
            permission_ids = [row['permission_id'] for row in cursor.fetchall()]
            return {"code": 200, "message": "success", "data": permission_ids}
    except Exception as e:
        logger.error(f"获取角色权限失败: {e}")
        return {"code": 500, "message": "获取角色权限失败"}

@app.post("/api/role-permissions/")
async def assign_role_permissions(assign_request: AssignPermissionRequest, db = Depends(get_db)):
    """为角色分配权限点"""
    try:
        with db.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("DELETE FROM role_permissions WHERE role_id = %s", (assign_request.role_id,))
            
            for permission_id in assign_request.permission_ids:
                cursor.execute(
                    "INSERT INTO role_permissions (role_id, permission_id) VALUES (%s, %s)",
                    (assign_request.role_id, permission_id)
                )
            db.commit()
            return {"code": 200, "message": "权限分配成功"}
    except Exception as e:
        logger.error(f"权限分配失败: {e}")
        return {"code": 500, "message": "权限分配失败"}

# 获取用户权限点列表
@app.get("/api/user/permissions")
async def get_user_permissions(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """获取当前用户的权限点列表"""
    try:
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 从token获取用户名
                token = credentials.credentials
                payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                username = payload.get("sub")
                
                if not username:
                    return {"code": 401, "message": "无效的访问令牌"}
                
                # 获取用户角色
                cursor.execute("SELECT r.id FROM users u JOIN user_roles ur ON u.id = ur.user_id JOIN roles r ON ur.role_id = r.id WHERE u.username = %s AND r.status = 1", (username,))
                roles = cursor.fetchall()
                
                if not roles:
                    return {"code": 200, "message": "success", "data": []}
                
                role_ids = [role['id'] for role in roles]
                
                # 获取角色权限（使用新的细粒度权限系统）
                placeholders = ','.join(['%s'] * len(role_ids))
                cursor.execute(f"""
                    SELECT DISTINCT p.code 
                    FROM permissions p 
                    JOIN role_permissions rp ON p.id = rp.permission_id 
                    WHERE rp.role_id IN ({placeholders}) AND p.status = 1
                """, role_ids)
                permissions = [row['code'] for row in cursor.fetchall()]
                
                return {"code": 200, "message": "success", "data": permissions}
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"获取用户权限失败: {e}")
        import traceback
        traceback.print_exc()
        return {"code": 500, "message": "获取用户权限失败"}



# 拆分匹配相关请求模型
class ExecuteMatchRequest(BaseModel):
    table_name: Optional[str] = None
    records: Optional[List[dict]] = None

class UpdateMatchRequest(BaseModel):
    table_name: Optional[str] = None
    row_id: str
    data: dict

class TableDataResponse(BaseModel):
    total: int
    page: int
    page_size: int
    columns: List[str]
    data: List[dict]

class MatchResultResponse(BaseModel):
    matched_count: int
    unmatched_count: int
    total: int

# 初始化拆分匹配服务
split_match_service = SplitMatchService(config)

@app.get("/api/split-match/databases/")
async def get_available_databases_list():
    """获取所有可用的数据库名称"""
    try:
        databases = get_available_databases(config)
        return databases
    except Exception as e:
        logger.error(f"获取数据库列表失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "message": "获取数据库列表失败"})

@app.get("/api/split-match/tables/")
async def get_yc_tables():
    """获取check_data数据库中所有以_yc结尾的表"""
    try:
        tables = split_match_service.get_yc_tables()
        return {"code": 200, "message": "success", "data": tables}
    except Exception as e:
        logger.error(f"获取_yc表列表失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "message": "获取表列表失败"})

@app.get("/api/split-match/data/")
async def get_table_data(
    table_name: Optional[str] = Query(None, description="表名"),
    filters: Optional[str] = Query(None, description="字段筛选条件"),
    page: int = Query(1, description="页码"),
    page_size: int = Query(20, description="每页条数")
):
    """获取指定表的数据，支持分页和字段筛选"""
    try:
        # 如果没有提供表名，从详单查询配置中读取（因为拆分匹配CF表与详单查询使用同一个表）
        if not table_name:
            table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else None
            if not table_name:
                return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置详单查询数据表"})
        
        # 解析筛选条件
        filters_dict = None
        if filters:
            import json
            try:
                filters_dict = json.loads(filters)
            except:
                pass
        
        result = split_match_service.get_table_data(
            table_name=table_name,
            filters=filters_dict,
            page=page,
            page_size=page_size
        )
        response = {"code": 200, "message": "success", "data": result}
        if 'debug' in result:
            response['debug'] = result['debug']
        return response
    except Exception as e:
        logger.error(f"获取表数据失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "message": "获取表数据失败"})

@app.get("/api/split-match/export/")
async def export_table_data(
    table_name: Optional[str] = Query(None, description="表名"),
    filters: Optional[str] = Query(None, description="字段筛选条件")
):
    """获取指定表的完整数据用于导出，支持字段筛选但不进行分页"""
    try:
        # 如果没有提供表名，从详单查询配置中读取
        if not table_name:
            table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else None
            if not table_name:
                return JSONResponse(status_code=400, content={"code": 400, "message": "请先选择数据表"})
        
        # 解析筛选条件
        filters_dict = None
        if filters:
            import json
            try:
                filters_dict = json.loads(filters)
            except:
                pass
        
        result = split_match_service.get_all_table_data(
            table_name=table_name,
            filters=filters_dict
        )
        return {"code": 200, "message": "success", "data": result}
    except Exception as e:
        logger.error(f"获取完整表数据失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": "获取完整表数据失败"})

@app.post("/api/split-match/execute/")
async def execute_match(request: ExecuteMatchRequest):
    """执行拆分匹配，根据通行标识ID进行匹配并更新备注字段"""
    try:
        # 如果没有提供表名，从详单查询配置中读取（因为拆分匹配CF表与详单查询使用同一个表）
        table_name = request.table_name
        if not table_name:
            table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else None
            if not table_name:
                return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置详单查询数据表"})
        
        # 如果有records参数，传给服务层；否则服务层会自己查询
        records = request.records
        
        result = split_match_service.execute_match(table_name, records)
        response = {
            "code": 200, 
            "message": "匹配完成", 
            "data": result
        }
        if 'debug' in result:
            response['debug'] = result['debug']
        return response
    except Exception as e:
        logger.error(f"执行匹配失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "message": f"执行匹配失败: {str(e)}"})

@app.post("/api/split-match/preview/")
async def preview_match(request: ExecuteMatchRequest):
    """预览执行匹配将要执行的SQL语句（不实际执行）"""
    try:
        table_name = request.table_name
        if not table_name:
            table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else None
            if not table_name:
                return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置详单查询数据表"})
        
        records = request.records
        result = split_match_service.preview_match(table_name, records)
        return {"code": 200, "message": "success", "data": result}
    except Exception as e:
        logger.error(f"预览SQL失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"预览SQL失败: {str(e)}"})

@app.post("/api/split-match/update/")
async def update_match_data(request: UpdateMatchRequest):
    """更新拆分匹配数据"""
    try:
        # 如果没有提供表名，从详单查询配置中读取（因为拆分匹配CF表与详单查询使用同一个表）
        table_name = request.table_name
        if not table_name:
            table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else None
            if not table_name:
                return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置详单查询数据表"})
        
        logger.info(f"接收到更新请求: table_name={table_name}, row_id={request.row_id}")
        logger.info(f"更新数据: {request.data}")
        
        if not table_name or not request.row_id or not request.data:
            logger.warning("缺少必要参数")
            return JSONResponse(status_code=400, content={"code": 400, "message": "缺少必要参数"})
        
        logger.info(f"调用 update_table_data 方法")
        success = split_match_service.update_table_data(table_name, request.row_id, request.data)
        logger.info(f"update_table_data 方法返回: {success}")
        
        if success:
            logger.info("更新成功")
            return {
                "code": 200, 
                "message": "更新成功", 
                "data": {"updated": True}
            }
        else:
            logger.warning("更新失败，没有行被更新")
            return {
                "code": 400, 
                "message": "更新失败", 
                "data": {"updated": False}
            }
    except Exception as e:
        logger.error(f"更新数据失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"更新数据失败: {str(e)}"})

@app.get("/api/split-match/cf-tables/")
async def get_cf_tables():
    """获取拆分数据表列表"""
    try:
        tables = split_match_service.search_cf_tables()
        return {"code": 200, "message": "success", "data": tables}
    except Exception as e:
        logger.error(f"获取cf表列表失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "message": "获取表列表失败"})

# 详单查询API
class DetailQueryRequest(BaseModel):
    pass_id: Optional[str] = None
    entry_name: Optional[str] = None
    exit_name: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    plate_number: Optional[str] = None
    vehicle_type: Optional[str] = None
    billing_method: Optional[str] = None
    medium: Optional[str] = None
    settlement_date: Optional[str] = None
    data_type: Optional[str] = None
    split_month: Optional[str] = None
    page: int = 1
    page_size: int = 20

@app.post("/api/detail-query/")
async def query_detail_data(request: DetailQueryRequest):
    """详单查询，支持多条件组合查询
    
    现有索引分析：
    ┌────────────────┬─────────────────────────────────────┬─────────┐
    │ 索引名         │ 字段                                │ 基数    │
    ├────────────────┼─────────────────────────────────────┼─────────┤
    │ index_a        │ 通行标识ID                          │ 5224万  │
    │ index_b        │ 实际车辆车牌号码+颜色               │ 829万   │
    │ index_c        │ 计费交易终点时间                    │ 4607万  │
    │ idx_trade_time │ 计费交易起点时间,计费交易终点时间   │ 5224万  │
    └────────────────┴─────────────────────────────────────┴─────────┘
    
    数据库优化建议（优先执行）：
    为提升查询性能，建议添加以下索引：
    
    1. 【最高优先级】拆分月份查询优化（经常使用）：
       CREATE INDEX idx_split_month ON `{table_name}`(`拆分月份`);
       
    2. 【高优先级】拆分月份+车牌号码复合索引（常用组合查询）：
       CREATE INDEX idx_split_month_plate ON `{table_name}`(`拆分月份`, `实际车辆车牌号码+颜色`);
       
    3. 【中优先级】拆分月份+时间范围复合索引：
       CREATE INDEX idx_split_month_time ON `{table_name}`(`拆分月份`, `计费交易起点时间`);
       
    4. 【可选】入口/出口名称查询优化：
       CREATE INDEX idx_entry ON `{table_name}`(`收费入口名称`);
       CREATE INDEX idx_exit ON `{table_name}`(`收费出口名称`);
    
    执行顺序建议：
    1. 先创建 idx_split_month（最快见效）
    2. 再创建 idx_split_month_plate（最常用的组合查询）
    3. 根据实际查询情况添加其他索引
    
    注意：
    - 创建索引会增加存储空间和写入操作的开销
    - 建议在业务低峰期执行索引创建
    - 创建索引后使用 EXPLAIN 分析查询执行计划确认生效
    """
    import asyncio
    import time
    from concurrent.futures import ThreadPoolExecutor
    
    # 从配置中获取表名
    table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else "202005-202311_cf_1215"
    if not table_name:
        return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置详单查询数据表"})
    
    query_start_time = time.time()
    logger.info(f"========== 详单查询开始 ==========")
    logger.info(f"查询参数: {request}")
    
    conn = None
    try:
        conn = create_db_connection("CHECK_DATA_DB", config)
        if not conn:
            logger.error("无法连接到数据库")
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到数据库"})
        
        def format_sql(sql: str, params_list: list) -> str:
            """格式化 SQL 语句，替换参数为实际值"""
            result = sql
            for param in params_list:
                if isinstance(param, str):
                    result = result.replace("%s", f"'{param}'", 1)
                else:
                    result = result.replace("%s", str(param), 1)
            return result
        
        def execute_query():
            """在单独的线程中执行查询"""
            step_start_time = time.time()
            
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                where_conditions = []
                params = []
                
                if request.pass_id:
                    where_conditions.append("`通行标识ID` LIKE %s")
                    params.append(f"{request.pass_id}")
                if request.entry_name:
                    where_conditions.append("`收费入口名称` LIKE %s")
                    params.append(f"{request.entry_name}%")
                if request.exit_name:
                    where_conditions.append("`收费出口名称` LIKE %s")
                    params.append(f"{request.exit_name}%")
                if request.start_time:
                    where_conditions.append("`计费交易起点时间` >= %s")
                    params.append(request.start_time)
                if request.end_time:
                    where_conditions.append("`计费交易终点时间` <= %s")
                    params.append(request.end_time)
                if request.plate_number:
                    where_conditions.append("`实际车辆车牌号码+颜色` LIKE %s")
                    params.append(f"{request.plate_number}%")
                if request.vehicle_type:
                    where_conditions.append("`收费车型` LIKE %s")
                    params.append(f"{request.vehicle_type}")
                if request.billing_method:
                    where_conditions.append("`计费方式` LIKE %s")
                    params.append(f"{request.billing_method}")
                if request.medium:
                    where_conditions.append("`通行介质` LIKE %s")
                    params.append(f"{request.medium}")
                if request.settlement_date:
                    where_conditions.append("`清分日` LIKE %s")
                    params.append(f"{request.settlement_date}")
                if request.data_type:
                    where_conditions.append("`拆分类型/数据类型` LIKE %s")
                    params.append(f"{request.data_type}")
                if request.split_month:
                    where_conditions.append("`拆分月份` LIKE %s")
                    params.append(f"{request.split_month}")
                
                where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
                
                logger.info(f"[1/4] 构建 WHERE 条件耗时: {time.time() - step_start_time:.3f}s")
                step_start_time = time.time()
                
                count_sql = f"SELECT COUNT(*) as total FROM `{table_name}` WHERE {where_clause}"
                formatted_count_sql = format_sql(count_sql, params)
                logger.info("=" * 100)
                logger.info(f"[2/4] ========== COUNT SQL 开始 ==========")
                logger.info(f"[2/4] {formatted_count_sql}")
                logger.info(f"[2/4] ========== COUNT SQL 结束 ==========")
                
                cursor.execute(count_sql, params)
                total_result = cursor.fetchone()
                total = total_result['total'] if total_result else 0
                count_duration = time.time() - step_start_time
                logger.info(f"[2/4] COUNT 查询完成，共 {total} 条记录，耗时: {count_duration:.3f}s")
                
                if count_duration > 1:
                    logger.warning(f"[提示] COUNT 查询耗时较长({count_duration:.3f}s)，建议添加索引优化！")
                    logger.warning(f"[提示] 推荐索引: CREATE INDEX idx_split_month ON `{table_name}`(`拆分月份`);")
                
                step_start_time = time.time()
                
                offset = (request.page - 1) * request.page_size
                select_sql = f"SELECT * FROM `{table_name}` WHERE {where_clause} LIMIT %s OFFSET %s"
                select_params = params + [request.page_size, offset]
                formatted_select_sql = format_sql(select_sql, select_params)
                logger.info(f"[3/4] ========== SELECT SQL 开始 ==========")
                logger.info(f"[3/4] {formatted_select_sql}")
                logger.info(f"[3/4] ========== SELECT SQL 结束 ==========")
                logger.info("=" * 100)
                
                cursor.execute(select_sql, select_params)
                data = cursor.fetchall()
                logger.info(f"[3/4] SELECT 查询完成，获取 {len(data)} 条记录，耗时: {time.time() - step_start_time:.3f}s")
                
                return total, data, formatted_count_sql, formatted_select_sql, count_duration
        
        loop = asyncio.get_event_loop()
        executor = ThreadPoolExecutor(max_workers=1)
        
        try:
            logger.info("[4/4] 开始执行数据库查询...")
            total, data, count_sql_str, select_sql_str, count_duration_val = await asyncio.wait_for(
                loop.run_in_executor(executor, execute_query),
                timeout=300
            )
            
            total_time = time.time() - query_start_time
            logger.info(f"========== 详单查询完成 ==========")
            logger.info(f"总耗时: {total_time:.3f}s, 共 {total} 条记录, 返回 {len(data)} 条")
            
            return {
                "code": 200, 
                "message": "success", 
                "data": {
                    "list": data,
                    "total": total,
                    "page": request.page,
                    "page_size": request.page_size
                },
                "debug": {
                    "count_sql": count_sql_str,
                    "select_sql": select_sql_str,
                    "total_time": total_time,
                    "count_duration": count_duration_val
                }
            }
        except asyncio.TimeoutError:
            total_time = time.time() - query_start_time
            logger.error(f"详单查询超时！总耗时: {total_time:.3f}s")
            return JSONResponse(
                status_code=408, 
                content={
                    "code": 408, 
                    "message": "查询超时，请减少查询范围或使用更精确的查询条件"
                }
            )
    except Exception as e:
        total_time = time.time() - query_start_time
        logger.error(f"详单查询失败！总耗时: {total_time:.3f}s, 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500, 
            content={
                "code": 500, 
                "message": f"查询失败: {str(e)}"
            }
        )
    finally:
        if conn:
            conn.close()

@app.get("/api/split-match/verify-pass-id/")
async def verify_pass_id(
    pass_id: str = Query(..., description="通行标识ID或核查通行标识"),
    verify_type: str = Query(..., description="核查类型：通行标识ID 或 核查通行标识"),
    user_id: Optional[int] = Query(None, description="操作用户ID"),
    username: Optional[str] = Query(None, description="操作用户名")
):
    """核查通行标识ID是否存在于DETAIL_QUERY表中
    
    返回：
    - exists: 是否存在
    - match_count: 匹配记录数
    - records: 前3条匹配记录的关键信息
    """
    import time

    query_start_time = time.time()
    logger.info(f"========== 核查通行标识开始 ==========")
    logger.info(f"核查ID: {pass_id}, 类型: {verify_type}, 用户: {username}")
    
    if not pass_id or not pass_id.strip():
        return JSONResponse(status_code=400, content={"code": 400, "message": "通行标识ID不能为空"})
    
    pass_id = pass_id.strip()
    
    table_name = config.get('DETAIL_QUERY', 'table_name') if config.has_section('DETAIL_QUERY') and config.has_option('DETAIL_QUERY', 'table_name') else None
    if not table_name:
        return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置详单查询数据表"})
    
    conn = None
    try:
        conn = create_db_connection("CHECK_DATA_DB", config)
        if not conn:
            logger.error("无法连接到数据库")
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到数据库"})
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            count_sql = f"SELECT COUNT(*) as total FROM `{table_name}` WHERE `通行标识ID` = %s"
            cursor.execute(count_sql, (pass_id,))
            count_result = cursor.fetchone()
            match_count = count_result['total'] if count_result else 0
            
            records = []
            if match_count > 0:
                select_sql = f"""
                    SELECT 
                        `通行标识ID`,
                        `实际车辆车牌号码+颜色`,
                        `收费入口名称`,
                        `收费出口名称`,
                        `计费交易起点时间`,
                        `计费交易终点时间`,
                        `收费车型`,
                        `通行介质`,
                        `计费方式`,
                        `清分日`
                    FROM `{table_name}` 
                    WHERE `通行标识ID` = %s 
                    LIMIT 3
                """
                cursor.execute(select_sql, (pass_id,))
                records = cursor.fetchall()
                
                for record in records:
                    for key, value in record.items():
                        if hasattr(value, 'isoformat'):
                            record[key] = value.isoformat()
            
            total_time = time.time() - query_start_time
            logger.info(f"========== 核查完成 ==========")
            logger.info(f"核查ID: {pass_id}, 匹配记录数: {match_count}, 耗时: {total_time:.3f}s")
            
            verify_result = {
                "exists": match_count > 0,
                "match_count": match_count,
                "records": records,
                "query_time": total_time
            }
            
            try:
                history_conn = create_db_connection("LOCAL_DB", config)
                if history_conn:
                    try:
                        with history_conn.cursor() as history_cursor:
                            insert_sql = """
                                INSERT INTO verify_history 
                                (pass_id, verify_type, table_name, match_count, verify_result, user_id, username, verify_time)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """
                            history_cursor.execute(insert_sql, (
                                pass_id,
                                verify_type,
                                table_name,
                                match_count,
                                json.dumps(verify_result, ensure_ascii=False),
                                user_id,
                                username,
                                datetime.now()
                            ))
                            history_conn.commit()
                            logger.info(f"核查历史记录已保存: {pass_id}")
                    except Exception as e:
                        logger.error(f"保存核查历史记录失败: {e}", exc_info=True)
                    finally:
                        history_conn.close()
            except Exception as e:
                logger.error(f"连接历史记录数据库失败: {e}", exc_info=True)
            
            return {
                "code": 200,
                "message": "success",
                "data": verify_result
            }
    
    except Exception as e:
        total_time = time.time() - query_start_time
        logger.error(f"核查失败！总耗时: {total_time:.3f}s, 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "code": 500,
                "message": f"核查失败: {str(e)}"
            }
        )
    finally:
        if conn:
            conn.close()

@app.get("/api/split-match/verify-history/")
async def get_verify_history(
    pass_id: Optional[str] = Query(None, description="通行标识ID"),
    user_id: Optional[int] = Query(None, description="用户ID"),
    page: int = Query(1, description="页码"),
    page_size: int = Query(20, description="每页条数")
):
    """获取核查历史记录"""
    try:
        conn = create_db_connection("LOCAL_DB", config)
        if not conn:
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到数据库"})
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            where_conditions = []
            params = []
            
            if pass_id:
                where_conditions.append("pass_id LIKE %s")
                params.append(f"%{pass_id}%")
            
            if user_id:
                where_conditions.append("user_id = %s")
                params.append(user_id)
            
            where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
            
            count_sql = f"SELECT COUNT(*) as total FROM verify_history WHERE {where_clause}"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']
            
            offset = (page - 1) * page_size
            select_sql = f"""
                SELECT 
                    id, pass_id, verify_type, table_name, match_count, 
                    verify_result, user_id, username, verify_time
                FROM verify_history 
                WHERE {where_clause}
                ORDER BY verify_time DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(select_sql, params + [page_size, offset])
            records = cursor.fetchall()
            
            for record in records:
                if record['verify_result']:
                    try:
                        record['verify_result'] = json.loads(record['verify_result'])
                    except:
                        pass
                if record['verify_time']:
                    record['verify_time'] = record['verify_time'].isoformat()
            
            return {
                "code": 200,
                "message": "success",
                "data": {
                    "list": records,
                    "total": total,
                    "page": page,
                    "page_size": page_size
                }
            }
    
    except Exception as e:
        logger.error(f"获取核查历史失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"获取核查历史失败: {str(e)}"})
    finally:
        if conn:
            conn.close()

class PathMatchRequest(BaseModel):
    timeRange: List[str]
    transactionTimeRange: List[str]
    includeUnits: str
    endUnit: str
    excludeExits: str
    excludeUnits: str
    excludeGreenChannel: bool = True
    vehicleTypes: str = ""

@app.get("/api/path-match/tables/")
async def get_path_match_tables():
    """获取路径匹配可用的表列表"""
    try:
        check_data_config = get_database_config(config, 'CHECK_DATA_DB')
        
        conn = create_db_connection('CHECK_DATA_DB', config)
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                query = """
                    SELECT TABLE_NAME 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_SCHEMA = %s 
                    ORDER BY TABLE_NAME
                """
                cursor.execute(query, (check_data_config['database'],))
                tables = [row['TABLE_NAME'] for row in cursor.fetchall()]
            
            return {"code": 200, "message": "success", "data": tables}
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"获取表列表失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"获取表列表失败: {str(e)}"})

class ProvinceQueryRequest(BaseModel):
    codes: List[str]

@app.post("/api/path-match/provinces")
async def get_provinces_by_codes(request: ProvinceQueryRequest):
    """根据收费站编码前14位查询省份"""
    try:
        if not request.codes:
            return {"code": 200, "message": "success", "data": {}}
        
        unique_codes = list(set(code[:14] for code in request.codes if code and len(code) >= 14))
        
        if not unique_codes:
            return {"code": 200, "message": "success", "data": {}}
        
        conn = create_db_connection('YIN_WU_DB', config)
        if not conn:
            logger.error("无法连接到yin_wu数据库")
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到yin_wu数据库"})
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                placeholders = ','.join(['%s'] * len(unique_codes))
                query = f"""
                    SELECT 收费站编码, 省份名称 
                    FROM station_prov 
                    WHERE 收费站编码 IN ({placeholders})
                """
                cursor.execute(query, unique_codes)
                rows = cursor.fetchall()
                
                code_province_map = {row['收费站编码']: row['省份名称'] for row in rows}
                
                return {"code": 200, "message": "success", "data": code_province_map}
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"查询省份失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"查询省份失败: {str(e)}"})

class PathDetailRequest(BaseModel):
    passageId: str
    tableName: str

@app.post("/api/path-match/detail")
async def get_path_match_detail(request: PathDetailRequest):
    """获取路径匹配单条记录详情"""
    try:
        table_name = request.tableName or config.get('PATH_MATCH', 'table_name') if config.has_section('PATH_MATCH') and config.has_option('PATH_MATCH', 'table_name') else None
        if not table_name:
            return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置路径匹配数据表"})
        
        conn = create_db_connection('CHECK_DATA_DB', config)
        if not conn:
            logger.error("无法连接到数据库")
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到数据库"})
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                sql = f"SELECT * FROM `{table_name}` WHERE `通行标识ID` = %s LIMIT 1"
                cursor.execute(sql, (request.passageId,))
                result = cursor.fetchone()
                
                if not result:
                    return JSONResponse(status_code=404, content={"code": 404, "message": "未找到该记录"})
                
                return {"code": 200, "message": "success", "data": result}
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"获取详情失败: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"获取详情失败: {str(e)}"})

@app.post("/api/path-match/search")
async def path_match_search(request: PathMatchRequest):
    """路径匹配搜索接口"""
    import time
    try:
        # 从配置中获取表名
        table_name = config.get('PATH_MATCH', 'table_name') if config.has_section('PATH_MATCH') and config.has_option('PATH_MATCH', 'table_name') else None
        if not table_name:
            return JSONResponse(status_code=400, content={"code": 400, "message": "请先在参数配置中设置路径匹配数据表"})
        
        query_start_time = time.time()
        timeRange = request.timeRange
        transactionTimeRange = request.transactionTimeRange
        includeUnits = request.includeUnits
        endUnit = request.endUnit
        excludeExits = request.excludeExits
        excludeUnits = request.excludeUnits
        excludeGreenChannel = request.excludeGreenChannel
        vehicleTypes = request.vehicleTypes
        
        all_vehicle_types = ['1', '2', '3', '4', '11', '12', '13', '14', '15', '16', '21', '22', '23', '24', '25', '26']
        
        logger.info(f"========== 路径匹配搜索开始 ==========")
        logger.info(f"查询参数: {request}")
        
        check_data_config = get_database_config(config, 'CHECK_DATA_DB')
        
        conn = create_db_connection('CHECK_DATA_DB', config)
        if not conn:
            logger.error("无法连接到数据库")
            return JSONResponse(status_code=500, content={"code": 500, "message": "无法连接到数据库"})
        
        def format_sql(sql_query: str, sql_params: list) -> str:
            """格式化 SQL 语句，替换参数为实际值"""
            # 先将 %% 替换为占位符，避免干扰参数替换
            temp_sql = sql_query.replace('%%', '__PERCENT_SIGN__')
            
            # 替换参数
            for param in sql_params:
                if isinstance(param, str):
                    temp_sql = temp_sql.replace('%s', f"'{param}'", 1)
                else:
                    temp_sql = temp_sql.replace('%s', str(param), 1)
            
            # 还原 %% 为 %
            debug_sql = temp_sql.replace('__PERCENT_SIGN__', '%')
            return debug_sql
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                sql = ""
                params = []
                
                select_fields = """`通行标识ID`, `收费入口编码`, `收费出口编码`, `收费入口名称`, `收费出口名称`, 
                    `计费交易起点时间`, `计费交易终点时间`, `实际车辆车牌号码+颜色`, `收费车型`, `清分日`, 
                    `拆分路段拆分金额`, `拆分收费单元名称组合`, `拆分收费时间组合`, `拆分类型/数据类型`"""
                
                sql += f"SELECT {select_fields} FROM `{{table_name}}` WHERE 1=1".format(table_name=table_name)
                
                if timeRange and len(timeRange) == 2:
                    startTime = timeRange[0]
                    endTime = timeRange[1]
                    
                    if transactionTimeRange and len(transactionTimeRange) == 2:
                        startDate = transactionTimeRange[0]
                        endDate = transactionTimeRange[1]
                        sql += " AND `清分日` >= %s "
                        sql += " AND `清分日` <= %s "
                        params.append(startDate)
                        params.append(endDate + ' 23:59:59')
                    else:
                        sql += " AND `计费交易起点时间` >= %s "
                        sql += " AND `计费交易终点时间` <= %s "
                        params.append(startTime.split(' ')[0])
                        params.append(endTime.split(' ')[0])
                    
                    if excludeGreenChannel:
                        sql += " AND `拆分类型/数据类型` != 36 "
                    
                    if vehicleTypes:
                        typesArray = [item.strip() for item in vehicleTypes.split(',') if item.strip()]
                        if len(typesArray) > 0 and len(typesArray) < len(all_vehicle_types):
                            placeholders = ', '.join(['%s'] * len(typesArray))
                            sql += f" AND `收费车型` IN ({placeholders}) "
                            params.extend(typesArray)
                    
                    if includeUnits and endUnit:
                        sql += " AND `拆分收费单元名称组合` LIKE %s "
                        params.append(f"%{includeUnits.strip()}%{endUnit.strip()}%")
                    elif includeUnits:
                        sql += " AND `拆分收费单元名称组合` LIKE %s "
                        params.append(f"%{includeUnits.strip()}%")
                    elif endUnit:
                        sql += " AND `拆分收费单元名称组合` LIKE %s "
                        params.append(f"%{endUnit.strip()}%")
                    
                    if excludeUnits:
                        excludeUnitsArray = [item.strip() for item in excludeUnits.split(',')]
                        for excludeUnit in excludeUnitsArray:
                            if excludeUnit:
                                sql += " AND `拆分收费单元名称组合` NOT LIKE %s "
                                params.append(f"%{excludeUnit}%")
                    
                    if excludeExits:
                        excludeExitsArray = [item.strip() for item in excludeExits.split(',')]
                        for excludeExit in excludeExitsArray:
                            if excludeExit:
                                sql += " AND `收费出口名称` NOT LIKE %s "
                                params.append(f"{excludeExit}%")
                else:
                    if excludeGreenChannel:
                        sql += " AND `拆分类型/数据类型` != 36 "
                    
                    if vehicleTypes:
                        typesArray = [item.strip() for item in vehicleTypes.split(',') if item.strip()]
                        if len(typesArray) > 0 and len(typesArray) < len(all_vehicle_types):
                            placeholders = ', '.join(['%s'] * len(typesArray))
                            sql += f" AND `收费车型` IN ({placeholders}) "
                            params.extend(typesArray)
                    
                    if includeUnits and endUnit:
                        sql += " AND `拆分收费单元名称组合` LIKE %s"
                        params.append(f"%{includeUnits.strip()}%{endUnit.strip()}%")
                    elif includeUnits:
                        sql += " AND `拆分收费单元名称组合` LIKE %s"
                        params.append(f"%{includeUnits.strip()}%")
                    elif endUnit:
                        sql += " AND `拆分收费单元名称组合` LIKE %s"
                        params.append(f"%{endUnit.strip()}%")
                    
                    if excludeExits:
                        excludeExitsArray = [item.strip() for item in excludeExits.split(',')]
                        for excludeExit in excludeExitsArray:
                            if excludeExit:
                                sql += " AND `收费出口名称` NOT LIKE %s"
                                params.append(f"{excludeExit}%")
                    
                    if excludeUnits:
                        excludeUnitsArray = [item.strip() for item in excludeUnits.split(',')]
                        for excludeUnit in excludeUnitsArray:
                            if excludeUnit:
                                sql += " AND `拆分收费单元名称组合` NOT LIKE %s"
                                params.append(f"%{excludeUnit}%")
                
                formatted_sql = format_sql(sql, params)
                logger.info("=" * 100)
                logger.info(f"========== SELECT SQL 开始 ==========")
                logger.info(f"{formatted_sql}")
                logger.info(f"========== SELECT SQL 结束 ==========")
                logger.info("=" * 100)
                
                step_start_time = time.time()
                cursor.execute(sql, params)
                results = cursor.fetchall()
                select_duration = time.time() - step_start_time
                
                # 第二步：在Python中过滤门架经过时间范围
                filtered_results = []
                if timeRange and len(timeRange) == 2:
                    target_start = datetime.strptime(timeRange[0], '%Y-%m-%d %H:%M:%S')
                    target_end = datetime.strptime(timeRange[1], '%Y-%m-%d %H:%M:%S')
                    
                    for item in results:
                        time_combination = item.get('拆分收费时间组合', '')
                        if not time_combination:
                            continue
                        
                        # 拆分时间组合
                        time_list = str(time_combination).split('|')
                        has_match = False
                        
                        for time_str in time_list:
                            time_str = time_str.strip()
                            if not time_str:
                                continue
                            
                            try:
                                # 解析时间
                                record_time = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
                                # 检查是否在目标时间范围内
                                if target_start <= record_time <= target_end:
                                    has_match = True
                                    break
                            except ValueError:
                                continue
                        
                        if has_match:
                            filtered_results.append(item)
                else:
                    filtered_results = results
                
                processedResults = []
                for index, item in enumerate(filtered_results):
                    result = dict(item)
                    processedResults.append(result)
                
                total_time = time.time() - query_start_time
                logger.info(f"========== 路径匹配搜索完成 ==========")
                logger.info(f"总耗时: {total_time:.3f}s, 查询耗时: {select_duration:.3f}s, 数据库返回 {len(results)} 条, 时间过滤后 {len(processedResults)} 条记录")
                
                return {
                    "code": 200, 
                    "message": "success", 
                    "data": processedResults, 
                    "count": len(processedResults),
                    "debug": {
                        "select_sql": formatted_sql,
                        "count_sql": formatted_sql,
                        "total_time": total_time,
                        "count_duration": select_duration
                    }
                }
                
        finally:
            conn.close()
            
    except Exception as e:
        total_time = time.time() - query_start_time
        logger.error(f"路径匹配搜索失败！总耗时: {total_time:.3f}s, 错误: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"code": 500, "message": f"搜索失败: {str(e)}"})

@app.get("/api/params-config/")
async def get_params_config():
    """获取参数配置"""
    result = {}
    for section in config.sections():
        result[section] = {}
        for key, value in config.items(section):
            result[section][key] = value
    return result

@app.post("/api/params-config/")
async def save_params_config(config_data: dict):
    """保存参数配置"""
    global config
    # 更新配置对象
    for section, section_data in config_data.items():
        if not config.has_section(section):
            config.add_section(section)
        for key, value in section_data.items():
            config.set(section, key, str(value))
    
    # 保存到文件
    if save_config(config):
        # 广播配置变更通知
        config_message = json.dumps({
            "type": "config_update",
            "timestamp": datetime.now().isoformat(),
            "message": "配置已更新",
            "config": {
                section: dict(config[section]) for section in config.sections()
            }
        })
        await manager.broadcast(config_message)
        return {"message": "配置保存成功"}
    else:
        return JSONResponse(status_code=500, content={"message": "配置保存失败"})

@app.on_event("startup")
async def startup_event():
    """应用启动时初始化连接池"""
    logger.info("应用启动，正在初始化数据库连接池...")
    try:
        common_db_types = ["USER_DB", "LOCAL_DB", "REMOTE_DB", "CHECK_DATA_DB"]
        for db_type in common_db_types:
            try:
                create_db_pool(db_type, config, max_connections=10)
            except Exception as e:
                logger.warning(f"预初始化 {db_type} 连接池失败: {e}，将在首次使用时创建")
        init_special_records_table()
        init_cloud_portal_account_table()
        logger.info("应用启动完成，数据库连接池已初始化")
    except Exception as e:
        logger.error(f"应用启动时初始化连接池失败: {e}", exc_info=True)

@app.get("/api/scheduled-tasks/")
async def get_scheduled_tasks(user: dict = Depends(require_permission('system:scheduled-tasks:view'))):
    """获取定时任务列表"""
    try:
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接数据库", "data": []}
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                cursor.execute("""
                    SELECT id, task_name, task_type, cron_expression, is_enabled,
                           last_run_time, next_run_time, last_run_status, last_run_message,
                           created_at, updated_at
                    FROM scheduled_tasks
                    ORDER BY id
                """)
                tasks = cursor.fetchall()
                
                for task in tasks:
                    if task['last_run_time']:
                        task['last_run_time'] = task['last_run_time'].strftime('%Y-%m-%d %H:%M:%S')
                    if task['next_run_time']:
                        task['next_run_time'] = task['next_run_time'].strftime('%Y-%m-%d %H:%M:%S')
                    if task['created_at']:
                        task['created_at'] = task['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                    if task['updated_at']:
                        task['updated_at'] = task['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
                
                return {"code": 200, "message": "success", "data": tasks}
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"获取定时任务列表失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}

@app.put("/api/scheduled-tasks/{task_id}")
async def update_scheduled_task(task_id: int, request_data: dict, user: dict = Depends(require_permission('system:scheduled-tasks:edit'))):
    """更新定时任务配置"""
    try:
        conn = create_db_connection("USER_DB", config) or create_db_connection("LOCAL_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接数据库"}
        
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                is_enabled = request_data.get('is_enabled')
                cron_expression = request_data.get('cron_expression')
                
                update_fields = []
                params = []
                
                if is_enabled is not None:
                    update_fields.append("is_enabled = %s")
                    params.append(1 if is_enabled else 0)
                
                if cron_expression:
                    update_fields.append("cron_expression = %s")
                    params.append(cron_expression)
                
                if not update_fields:
                    return {"code": 400, "message": "没有要更新的字段"}
                
                params.append(task_id)
                sql = f"UPDATE scheduled_tasks SET {', '.join(update_fields)} WHERE id = %s"
                cursor.execute(sql, params)
                conn.commit()
                
                return {"code": 200, "message": "更新成功"}
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"更新定时任务失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}

@app.post("/api/scheduled-tasks/{task_name}/run")
async def run_scheduled_task(task_name: str, user: dict = Depends(require_permission('system:scheduled-tasks:run'))):
    """手动执行定时任务"""
    history_id = None
    start_time = None
    
    try:
        start_time = datetime.now()
        history_id = start_task_execution(task_name)
        
        if task_name == 'dashboard_statistics_daily':
            result = run_statistics_task()
            
            status = 'success' if result['success'] else 'failed'
            update_task_status(task_name, status, result['message'])
            
            end_time = datetime.now()
            duration = int((end_time - start_time).total_seconds()) if start_time else None
            
            end_task_execution(
                history_id, 
                status, 
                result['message'], 
                {'data': result.get('data')}, 
                duration
            )
            
            return {"code": 200 if result['success'] else 500, "message": result['message'], "data": result}
        elif task_name == 'cloud_portal_data_sync':
            access_token = user.get('cloud_portal_token')
            if not access_token:
                if history_id:
                    end_task_execution(history_id, 'failed', '云门户未登录，请先登录云门户账号')
                return {"code": 401, "message": "云门户未登录，请先登录云门户账号"}
            
            portal_session_id = None
            try:
                conn = create_db_connection("USER_DB", config)
                if conn:
                    with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                        cursor.execute("SELECT portal_session_id FROM cloud_portal_accounts WHERE user_id = %s", (user['id'],))
                        account = cursor.fetchone()
                        if account:
                            portal_session_id = account.get('portal_session_id')
                    conn.close()
            except Exception as e:
                logger.warning(f"获取portal_session_id失败: {e}")
            
            result = run_cloud_portal_data_sync(access_token, portal_session_id)
            
            status = 'success' if result['success'] else 'failed'
            update_task_status(task_name, status, result['message'])
            
            end_time = datetime.now()
            duration = int((end_time - start_time).total_seconds()) if start_time else None
            
            end_task_execution(
                history_id, 
                status, 
                result['message'], 
                {'statistics': result.get('statistics')}, 
                duration
            )
            
            return {"code": 200 if result['success'] else 500, "message": result['message'], "data": result}
        else:
            if history_id:
                end_task_execution(history_id, 'failed', f"未知的任务: {task_name}")
            return {"code": 404, "message": f"未知的任务: {task_name}"}
    except Exception as e:
        logger.error(f"执行定时任务失败: {e}")
        update_task_status(task_name, 'failed', str(e))
        
        if history_id:
            from datetime import datetime
            end_time = datetime.now()
            duration = int((end_time - start_time).total_seconds()) if start_time else None
            end_task_execution(history_id, 'failed', str(e), None, duration)
        
        return {"code": 500, "message": f"执行失败: {str(e)}"}

@app.get("/api/dashboard-statistics/")
async def get_dashboard_stats():
    """获取Dashboard统计数据"""
    try:
        stats = get_dashboard_statistics()
        if stats:
            if stats.get('created_at'):
                stats['created_at'] = stats['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if stats.get('updated_at'):
                stats['updated_at'] = stats['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
        return {"code": 200, "message": "success", "data": stats}
    except Exception as e:
        logger.error(f"获取Dashboard统计数据失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": None}

@app.get("/api/task-execution-history/")
async def get_execution_history(task_name: str = None, limit: int = 20, user: dict = Depends(require_permission('system:scheduled-tasks:view'))):
    """获取任务执行历史"""
    try:
        history = get_task_execution_history(task_name, limit)
        return {"code": 200, "message": "success", "data": history}
    except Exception as e:
        logger.error(f"获取任务执行历史失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}

@app.get("/api/cloud-portal-data/")
async def get_cloud_portal_data_api(user: dict = Depends(require_permission('system:scheduled-tasks:view'))):
    """获取云门户基础数据"""
    try:
        data = get_cloud_portal_data()
        if data:
            return {"code": 200, "message": "success", "data": data}
        else:
            return {"code": 404, "message": "暂无数据，请先执行数据同步", "data": None}
    except Exception as e:
        logger.error(f"获取云门户数据失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": None}

class CloudPortalDataSyncRequest(BaseModel):
    access_token: str
    session_id: str = None

@app.post("/api/cloud-portal-data/sync")
async def sync_cloud_portal_data(request: CloudPortalDataSyncRequest, user: dict = Depends(require_permission('system:scheduled-tasks:run'))):
    """执行云门户数据同步"""
    history_id = None
    start_time = None
    
    try:
        start_time = datetime.now()
        history_id = start_task_execution('cloud_portal_data_sync')
        
        if not request.access_token:
            if history_id:
                end_task_execution(history_id, 'failed', '云门户未登录，请先登录云门户账号')
            return {"code": 401, "message": "云门户未登录，请先登录云门户账号"}
        
        result = run_cloud_portal_data_sync(request.access_token, request.session_id)
        
        status = 'success' if result['success'] else 'failed'
        update_task_status('cloud_portal_data_sync', status, result['message'])
        
        end_time = datetime.now()
        duration = int((end_time - start_time).total_seconds()) if start_time else None
        
        end_task_execution(
            history_id, 
            status, 
            result['message'], 
            {'statistics': result.get('statistics')}, 
            duration
        )
        
        return {"code": 200 if result['success'] else 500, "message": result['message'], "data": result}
    except Exception as e:
        logger.error(f"执行云门户数据同步失败: {e}")
        update_task_status('cloud_portal_data_sync', 'failed', str(e))
        
        if history_id:
            end_time = datetime.now()
            duration = int((end_time - start_time).total_seconds()) if start_time else None
            end_task_execution(history_id, 'failed', str(e), None, duration)
        
        return {"code": 500, "message": f"执行失败: {str(e)}"}

@app.get("/api/cloud-portal-data/status")
async def get_cloud_portal_data_status(user: dict = Depends(require_permission('system:scheduled-tasks:view'))):
    """获取云门户数据同步状态"""
    try:
        data = get_cloud_portal_data()
        if data:
            return {
                "code": 200, 
                "message": "success", 
                "data": {
                    "has_data": True,
                    "last_update": data.get('last_update'),
                    "total_centers": data.get('total_centers', 0),
                    "total_road_sections": data.get('total_road_sections', 0),
                    "total_gantries": data.get('total_gantries', 0)
                }
            }
        else:
            return {
                "code": 200,
                "message": "success",
                "data": {
                    "has_data": False,
                    "last_update": None,
                    "total_centers": 0,
                    "total_road_sections": 0,
                    "total_gantries": 0
                }
            }
    except Exception as e:
        logger.error(f"获取云门户数据状态失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": None}

@app.get("/api/stations/")
async def get_stations(keyword: str = None, limit: int = 50):
    """获取收费站列表（支持模糊搜索）"""
    try:
        conn = create_db_connection('YIN_WU_DB', config)
        if not conn:
            logger.error("无法连接到yin_wu数据库")
            return {"code": 500, "message": "无法连接到yin_wu数据库", "data": []}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            if keyword:
                cursor.execute("""
                    SELECT DISTINCT 收费站编码, 省份名称, 收费站名称
                    FROM station_prov 
                    WHERE 收费站编码 IS NOT NULL 
                      AND (收费站编码 LIKE %s OR 省份名称 LIKE %s OR 收费站名称 LIKE %s)
                    ORDER BY 省份名称
                    LIMIT %s
                """, (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', limit))
            else:
                cursor.execute("""
                    SELECT DISTINCT 收费站编码, 省份名称, 收费站名称
                    FROM station_prov 
                    WHERE 收费站编码 IS NOT NULL 
                    ORDER BY 省份名称
                    LIMIT %s
                """, (limit,))
            
            rows = cursor.fetchall()
            
            stations = []
            for row in rows:
                station_code = row['收费站编码']
                station_name = row.get('收费站名称') or ''
                province_name = row['省份名称'] or station_code
                label_parts = [station_name, station_code, province_name]
                label = ' - '.join([p for p in label_parts if p])
                stations.append({
                    "value": station_code,
                    "label": label
                })
            
            return {"code": 200, "message": "success", "data": stations}
    except Exception as e:
        logger.error(f"获取收费站列表失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}
    finally:
        if 'conn' in locals():
            conn.close()

class SpecialRecordCreate(BaseModel):
    record_type: str
    date: str
    exit_station: str = None
    entry_station: str = None
    lane_number: str = None
    transaction_time: str = None
    license_plate: str = None
    reason: str = None
    remark: str = None

class SpecialRecordUpdate(BaseModel):
    record_type: str = None
    date: str = None
    exit_station: str = None
    entry_station: str = None
    lane_number: str = None
    transaction_time: str = None
    license_plate: str = None
    reason: str = None
    remark: str = None

def init_special_records_table():
    """初始化特情记录表"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            logger.error("无法连接到数据库创建特情记录表")
            return False
        
        with conn.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS special_records (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    record_type VARCHAR(20) NOT NULL COMMENT '记录类型: u-car/unpaid-car/no-truck',
                    date DATE NOT NULL COMMENT '日期',
                    exit_station VARCHAR(100) COMMENT '收费出口',
                    entry_station VARCHAR(100) COMMENT '收费入口',
                    lane_number VARCHAR(20) COMMENT '车道号',
                    transaction_time TIME COMMENT '交易时间',
                    license_plate VARCHAR(20) COMMENT '车牌',
                    reason VARCHAR(500) COMMENT '原因',
                    remark VARCHAR(500) COMMENT '备注',
                    created_by INT COMMENT '创建人ID',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
                    INDEX idx_record_type (record_type),
                    INDEX idx_date (date),
                    INDEX idx_license_plate (license_plate)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='特情记录表'
            """)
            conn.commit()
            logger.info("特情记录表初始化成功")
            return True
    except Exception as e:
        logger.error(f"初始化特情记录表失败: {e}")
        return False
    finally:
        if 'conn' in locals():
            conn.close()

def init_cloud_portal_account_table():
    """初始化云门户账号绑定表"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            logger.error("无法连接到数据库创建云门户账号绑定表")
            return False
        
        with conn.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS cloud_portal_accounts (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL COMMENT '用户ID',
                    portal_username VARCHAR(100) NOT NULL COMMENT '云门户用户名',
                    portal_password VARCHAR(255) NOT NULL COMMENT '云门户密码(加密存储)',
                    portal_session_id VARCHAR(255) DEFAULT NULL COMMENT '云门户会话ID',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
                    UNIQUE KEY uk_user_id (user_id),
                    INDEX idx_portal_username (portal_username)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='云门户账号绑定表'
            """)
            try:
                cursor.execute("ALTER TABLE cloud_portal_accounts ADD COLUMN portal_session_id VARCHAR(255) DEFAULT NULL COMMENT '云门户会话ID'")
                conn.commit()
                logger.info("云门户账号绑定表添加portal_session_id字段成功")
            except Exception as alter_err:
                if "Duplicate column name" in str(alter_err):
                    logger.info("portal_session_id字段已存在，跳过添加")
                else:
                    logger.warning(f"添加portal_session_id字段时出现警告: {alter_err}")
            
            conn.commit()
            logger.info("云门户账号绑定表初始化成功")
            return True
    except Exception as e:
        logger.error(f"初始化云门户账号绑定表失败: {e}")
        return False
    finally:
        if 'conn' in locals():
            conn.close()

class CloudPortalAccountCreate(BaseModel):
    portal_username: str
    portal_password: str

class CloudPortalAccountUpdate(BaseModel):
    portal_username: str = None
    portal_password: str = None

@app.get("/api/cloud-portal-account/")
async def get_cloud_portal_account(user: dict = Depends(get_current_user)):
    """获取当前用户的云门户账号绑定"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": None}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT id, user_id, portal_username, created_at, updated_at
                FROM cloud_portal_accounts
                WHERE user_id = %s
            """, (user['id'],))
            account = cursor.fetchone()
            
            if account:
                return {
                    "code": 200,
                    "message": "获取成功",
                    "data": {
                        "id": account['id'],
                        "user_id": account['user_id'],
                        "portal_username": account['portal_username'],
                        "has_password": True,
                        "created_at": account['created_at'].strftime('%Y-%m-%d %H:%M:%S') if account['created_at'] else None,
                        "updated_at": account['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if account['updated_at'] else None
                    }
                }
            else:
                return {"code": 200, "message": "未绑定云门户账号", "data": None}
    except Exception as e:
        logger.error(f"获取云门户账号绑定失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": None}
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/cloud-portal-account/")
async def create_or_update_cloud_portal_account(
    request: CloudPortalAccountCreate,
    user: dict = Depends(get_current_user)
):
    """创建或更新云门户账号绑定"""
    try:
        if not request.portal_username or not request.portal_password:
            return {"code": 400, "message": "用户名和密码不能为空"}
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        encrypted_password = encrypt_password(request.portal_password)
        
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM cloud_portal_accounts WHERE user_id = %s
            """, (user['id'],))
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    UPDATE cloud_portal_accounts
                    SET portal_username = %s, portal_password = %s
                    WHERE user_id = %s
                """, (request.portal_username, encrypted_password, user['id']))
            else:
                cursor.execute("""
                    INSERT INTO cloud_portal_accounts (user_id, portal_username, portal_password)
                    VALUES (%s, %s, %s)
                """, (user['id'], request.portal_username, encrypted_password))
            
            conn.commit()
            return {"code": 200, "message": "绑定成功"}
    except Exception as e:
        logger.error(f"绑定云门户账号失败: {e}")
        return {"code": 500, "message": f"绑定失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/cloud-portal-account/")
async def delete_cloud_portal_account(user: dict = Depends(get_current_user)):
    """删除云门户账号绑定"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("""
                DELETE FROM cloud_portal_accounts WHERE user_id = %s
            """, (user['id'],))
            conn.commit()
            return {"code": 200, "message": "解绑成功"}
    except Exception as e:
        logger.error(f"解绑云门户账号失败: {e}")
        return {"code": 500, "message": f"解绑失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/cloud-portal-account/credentials")
async def get_cloud_portal_credentials(user: dict = Depends(get_current_user)):
    """获取当前用户的云门户账号凭证（用于自动登录）"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": None}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT portal_username, portal_password, portal_session_id
                FROM cloud_portal_accounts
                WHERE user_id = %s
            """, (user['id'],))
            account = cursor.fetchone()
            
            if account:
                decrypted_password = decrypt_password(account['portal_password'])
                return {
                    "code": 200,
                    "message": "获取成功",
                    "data": {
                        "portal_username": account['portal_username'],
                        "portal_password": decrypted_password,
                        "portal_session_id": account['portal_session_id'] or ''
                    }
                }
            else:
                return {"code": 200, "message": "未绑定云门户账号", "data": None}
    except Exception as e:
        logger.error(f"获取云门户凭证失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": None}
    finally:
        if 'conn' in locals():
            conn.close()

class CloudPortalSessionUpdate(BaseModel):
    session_id: str

@app.post("/api/cloud-portal-account/session")
async def update_cloud_portal_session(data: CloudPortalSessionUpdate, user: dict = Depends(get_current_user)):
    """更新云门户会话ID"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                UPDATE cloud_portal_accounts
                SET portal_session_id = %s
                WHERE user_id = %s
            """, (data.session_id if data.session_id else None, user['id']))
            conn.commit()
            
            return {"code": 200, "message": "会话ID更新成功"}
    except Exception as e:
        logger.error(f"更新云门户会话ID失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/special-records/")
async def get_special_records(
    record_type: str = None,
    date: str = None,
    license_plate: str = None,
    page: int = 1,
    page_size: int = 10,
    user: dict = Depends(require_permission('data-records:special-records:view'))
):
    """获取特情记录列表"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": []}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            where_clauses = []
            params = []
            
            if record_type:
                where_clauses.append("sr.record_type = %s")
                params.append(record_type)
            if date:
                where_clauses.append("sr.date = %s")
                params.append(date)
            if license_plate:
                where_clauses.append("sr.license_plate LIKE %s")
                params.append(f"%{license_plate}%")
            
            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            count_sql = f"SELECT COUNT(*) as total FROM special_records sr WHERE {where_sql}"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']
            
            offset = (page - 1) * page_size
            data_sql = f"""
                SELECT sr.id, sr.record_type, sr.date, sr.exit_station, sr.entry_station, 
                       sr.lane_number, sr.transaction_time, sr.license_plate, sr.reason, sr.remark,
                       sr.created_at, sr.updated_at,
                       es.收费站名称 as exit_station_name,
                       ens.收费站名称 as entry_station_name
                FROM special_records sr
                LEFT JOIN yin_wu.station_prov es ON sr.exit_station = es.收费站编码
                LEFT JOIN yin_wu.station_prov ens ON sr.entry_station = ens.收费站编码
                WHERE {where_sql}
                ORDER BY sr.created_at DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(data_sql, params + [page_size, offset])
            records = cursor.fetchall()
            
            for record in records:
                if record['date']:
                    record['date'] = record['date'].strftime('%Y-%m-%d')
                if record['transaction_time']:
                    if hasattr(record['transaction_time'], 'strftime'):
                        record['transaction_time'] = record['transaction_time'].strftime('%H:%M:%S')
                    else:
                        total_seconds = int(record['transaction_time'].total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        record['transaction_time'] = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                if record['created_at']:
                    record['created_at'] = record['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                if record['updated_at']:
                    record['updated_at'] = record['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
                record['entry_station_display'] = record.get('entry_station_name') or record.get('entry_station', '')
                record['exit_station_display'] = record.get('exit_station_name') or record.get('exit_station', '')
            
            return {
                "code": 200,
                "message": "success",
                "data": {
                    "list": records,
                    "total": total,
                    "page": page,
                    "pageSize": page_size
                }
            }
    except Exception as e:
        logger.error(f"获取特情记录失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/special-records/")
async def create_special_record(
    record: SpecialRecordCreate,
    user: dict = Depends(require_permission('data-records:special-records:create'))
):
    """创建特情记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO special_records 
                (record_type, date, exit_station, entry_station, lane_number, 
                 transaction_time, license_plate, reason, remark, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                record.record_type, record.date, record.exit_station, record.entry_station,
                record.lane_number, record.transaction_time, record.license_plate,
                record.reason, record.remark, user.get('id')
            ))
            conn.commit()
            
            return {"code": 200, "message": "添加成功", "data": {"id": cursor.lastrowid}}
    except Exception as e:
        logger.error(f"创建特情记录失败: {e}")
        return {"code": 500, "message": f"添加失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.put("/api/special-records/{record_id}")
async def update_special_record(
    record_id: int,
    record: SpecialRecordUpdate,
    user: dict = Depends(require_permission('data-records:special-records:edit'))
):
    """更新特情记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            update_fields = []
            params = []
            
            if record.record_type is not None:
                update_fields.append("record_type = %s")
                params.append(record.record_type)
            if record.date is not None:
                update_fields.append("date = %s")
                params.append(record.date)
            if record.exit_station is not None:
                update_fields.append("exit_station = %s")
                params.append(record.exit_station)
            if record.entry_station is not None:
                update_fields.append("entry_station = %s")
                params.append(record.entry_station)
            if record.lane_number is not None:
                update_fields.append("lane_number = %s")
                params.append(record.lane_number)
            if record.transaction_time is not None:
                update_fields.append("transaction_time = %s")
                params.append(record.transaction_time)
            if record.license_plate is not None:
                update_fields.append("license_plate = %s")
                params.append(record.license_plate)
            if record.reason is not None:
                update_fields.append("reason = %s")
                params.append(record.reason)
            if record.remark is not None:
                update_fields.append("remark = %s")
                params.append(record.remark)
            
            if not update_fields:
                return {"code": 400, "message": "没有需要更新的字段"}
            
            params.append(record_id)
            sql = f"UPDATE special_records SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(sql, params)
            conn.commit()
            
            if cursor.rowcount == 0:
                return {"code": 404, "message": "记录不存在"}
            
            return {"code": 200, "message": "更新成功"}
    except Exception as e:
        logger.error(f"更新特情记录失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/special-records/{record_id}")
async def delete_special_record(
    record_id: int,
    user: dict = Depends(require_permission('data-records:special-records:delete'))
):
    """删除特情记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM special_records WHERE id = %s", (record_id,))
            conn.commit()
            
            if cursor.rowcount == 0:
                return {"code": 404, "message": "记录不存在"}
            
            return {"code": 200, "message": "删除成功"}
    except Exception as e:
        logger.error(f"删除特情记录失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

class BatchDeleteRequest(BaseModel):
    ids: List[int]

@app.post("/api/special-records/batch-delete")
async def batch_delete_special_records(
    request: BatchDeleteRequest,
    user: dict = Depends(require_permission('data-records:special-records:delete'))
):
    """批量删除特情记录"""
    try:
        if not request.ids:
            return {"code": 400, "message": "请选择要删除的记录"}
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(request.ids))
            cursor.execute(f"DELETE FROM special_records WHERE id IN ({placeholders})", request.ids)
            conn.commit()
            
            return {"code": 200, "message": f"成功删除 {cursor.rowcount} 条记录"}
    except Exception as e:
        logger.error(f"批量删除特情记录失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/special-records/export/")
async def export_special_records(
    record_type: str = None,
    user: dict = Depends(require_permission('data-records:special-records:view'))
):
    """导出特情记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到数据库")
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            where_clause = "WHERE sr.record_type = %s" if record_type else ""
            params = [record_type] if record_type else []
            
            sql = f"""
                SELECT sr.id, sr.record_type, sr.date, sr.exit_station, sr.entry_station, 
                       sr.lane_number, sr.transaction_time, sr.license_plate, sr.reason, sr.remark,
                       sr.created_at, sr.updated_at,
                       es.收费站名称 as exit_station_name,
                       ens.收费站名称 as entry_station_name
                FROM special_records sr
                LEFT JOIN yin_wu.station_prov es ON sr.exit_station = es.收费站编码
                LEFT JOIN yin_wu.station_prov ens ON sr.entry_station = ens.收费站编码
                {where_clause}
                ORDER BY sr.created_at DESC
            """
            cursor.execute(sql, params)
            records = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "特情记录"
            
            headers = ["序号", "记录类型", "日期", "收费出口", "收费入口", "车道号", 
                      "交易时间", "车牌", "原因", "备注", "创建时间", "更新时间"]
            ws.append(headers)
            
            for idx, record in enumerate(records, 1):
                exit_station_display = record.get('exit_station_name') or record.get('exit_station', '')
                entry_station_display = record.get('entry_station_name') or record.get('entry_station', '')
                row = [
                    idx,
                    record.get('record_type', ''),
                    record['date'].strftime('%Y-%m-%d') if record.get('date') else '',
                    exit_station_display,
                    entry_station_display,
                    record.get('lane_number', ''),
                    format_time_value(record.get('transaction_time')),
                    record.get('license_plate', ''),
                    record.get('reason', ''),
                    record.get('remark', ''),
                    record['created_at'].strftime('%Y-%m-%d %H:%M:%S') if record.get('created_at') else '',
                    record['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if record.get('updated_at') else ''
                ]
                ws.append(row)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"特情记录_{record_type or 'all'}_{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib').parse.quote(filename)}"}
            )
    except Exception as e:
        logger.error(f"导出特情记录失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

class ExportByIdsRequest(BaseModel):
    ids: List[int]
    record_type: str = None

@app.post("/api/special-records/export/")
async def export_special_records_by_ids(
    request: ExportByIdsRequest,
    user: dict = Depends(require_permission('data-records:special-records:view'))
):
    """批量导出选中的特情记录"""
    try:
        if not request.ids:
            raise HTTPException(status_code=400, detail="请选择要导出的记录")
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到数据库")
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            placeholders = ','.join(['%s'] * len(request.ids))
            sql = f"""
                SELECT sr.id, sr.record_type, sr.date, sr.exit_station, sr.entry_station, 
                       sr.lane_number, sr.transaction_time, sr.license_plate, sr.reason, sr.remark,
                       sr.created_at, sr.updated_at,
                       es.收费站名称 as exit_station_name,
                       ens.收费站名称 as entry_station_name
                FROM special_records sr
                LEFT JOIN yin_wu.station_prov es ON sr.exit_station = es.收费站编码
                LEFT JOIN yin_wu.station_prov ens ON sr.entry_station = ens.收费站编码
                WHERE sr.id IN ({placeholders})
                ORDER BY sr.created_at DESC
            """
            cursor.execute(sql, request.ids)
            records = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "特情记录"
            
            headers = ["序号", "记录类型", "日期", "收费出口", "收费入口", "车道号", 
                      "交易时间", "车牌", "原因", "备注", "创建时间", "更新时间"]
            ws.append(headers)
            
            for idx, record in enumerate(records, 1):
                exit_station_display = record.get('exit_station_name') or record.get('exit_station', '')
                entry_station_display = record.get('entry_station_name') or record.get('entry_station', '')
                row = [
                    idx,
                    record.get('record_type', ''),
                    record['date'].strftime('%Y-%m-%d') if record.get('date') else '',
                    exit_station_display,
                    entry_station_display,
                    record.get('lane_number', ''),
                    format_time_value(record.get('transaction_time')),
                    record.get('license_plate', ''),
                    record.get('reason', ''),
                    record.get('remark', ''),
                    record['created_at'].strftime('%Y-%m-%d %H:%M:%S') if record.get('created_at') else '',
                    record['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if record.get('updated_at') else ''
                ]
                ws.append(row)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"特情记录_选中{len(records)}条_{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib').parse.quote(filename)}"}
            )
    except Exception as e:
        logger.error(f"批量导出特情记录失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/special-records/import/")
async def import_special_records(
    file: UploadFile = File(...),
    record_type: str = Form(...),
    user: dict = Depends(require_permission('data-records:special-records:create'))
):
    """导入特情记录"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            return {"code": 400, "message": "只支持Excel文件格式(.xlsx, .xls)"}
        
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return {"code": 400, "message": "Excel文件中没有数据"}
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        imported_count = 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or not any(row):
                    continue
                
                try:
                    date_val = row[2] if len(row) > 2 else None
                    if hasattr(date_val, 'strftime'):
                        date_val = date_val.strftime('%Y-%m-%d')
                    
                    time_val = row[6] if len(row) > 6 else None
                    if hasattr(time_val, 'strftime'):
                        time_val = time_val.strftime('%H:%M:%S')
                    
                    cursor.execute("""
                        INSERT INTO special_records 
                        (record_type, date, exit_station, entry_station, lane_number, 
                         transaction_time, license_plate, reason, remark, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        record_type,
                        date_val,
                        str(row[3]) if len(row) > 3 and row[3] else None,
                        str(row[4]) if len(row) > 4 and row[4] else None,
                        str(row[5]) if len(row) > 5 and row[5] else None,
                        time_val,
                        str(row[7]) if len(row) > 7 and row[7] else None,
                        str(row[8]) if len(row) > 8 and row[8] else None,
                        str(row[9]) if len(row) > 9 and row[9] else None,
                        user.get('id')
                    ))
                    imported_count += 1
                except Exception as e:
                    logger.warning(f"导入行失败: {e}, 行数据: {row}")
                    continue
            
            conn.commit()
        
        return {"code": 200, "message": "导入成功", "data": {"count": imported_count}}
    except Exception as e:
        logger.error(f"导入特情记录失败: {e}")
        return {"code": 500, "message": f"导入失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

class SchedulingGroupCreate(BaseModel):
    name: str
    description: str = None

class SchedulingGroupUpdate(BaseModel):
    name: str = None
    description: str = None

class SchedulingShiftCreate(BaseModel):
    name: str
    start_time: str
    end_time: str
    description: str = None

class SchedulingShiftUpdate(BaseModel):
    name: str = None
    start_time: str = None
    end_time: str = None
    description: str = None

class SchedulingStaffCreate(BaseModel):
    name: str
    employee_id: str = None
    group_id: int = None
    phone: str = None

class SchedulingStaffUpdate(BaseModel):
    name: str = None
    employee_id: str = None
    group_id: int = None
    phone: str = None
    status: int = None

class SchedulingRecordCreate(BaseModel):
    staff_id: int = None
    group_id: int
    shift_id: int
    schedule_date: str
    remark: str = None

class SchedulingRecordUpdate(BaseModel):
    staff_id: int = None
    group_id: int = None
    shift_id: int = None
    schedule_date: str = None
    remark: str = None

@app.get("/api/scheduling/groups/")
async def get_scheduling_groups(
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """获取班组列表"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": []}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM scheduling_groups ORDER BY id")
            groups = cursor.fetchall()
            return {"code": 200, "message": "success", "data": groups}
    except Exception as e:
        logger.error(f"获取班组列表失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/groups/")
async def create_scheduling_group(
    group: SchedulingGroupCreate,
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """创建班组"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO scheduling_groups (name, description) VALUES (%s, %s)",
                (group.name, group.description)
            )
            conn.commit()
            return {"code": 200, "message": "创建成功", "data": {"id": cursor.lastrowid}}
    except Exception as e:
        logger.error(f"创建班组失败: {e}")
        return {"code": 500, "message": f"创建失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.put("/api/scheduling/groups/{group_id}")
async def update_scheduling_group(
    group_id: int,
    group: SchedulingGroupUpdate,
    user: dict = Depends(require_permission('data-records:scheduling:edit'))
):
    """更新班组"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            update_fields = []
            params = []
            if group.name is not None:
                update_fields.append("name = %s")
                params.append(group.name)
            if group.description is not None:
                update_fields.append("description = %s")
                params.append(group.description)
            
            if not update_fields:
                return {"code": 400, "message": "没有需要更新的字段"}
            
            params.append(group_id)
            sql = f"UPDATE scheduling_groups SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(sql, params)
            conn.commit()
            return {"code": 200, "message": "更新成功"}
    except Exception as e:
        logger.error(f"更新班组失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/scheduling/groups/{group_id}")
async def delete_scheduling_group(
    group_id: int,
    user: dict = Depends(require_permission('data-records:scheduling:delete'))
):
    """删除班组"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM scheduling_groups WHERE id = %s", (group_id,))
            conn.commit()
            return {"code": 200, "message": "删除成功"}
    except Exception as e:
        logger.error(f"删除班组失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/groups/export/")
async def export_scheduling_groups(
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """导出班组数据"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到数据库")
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT name, description, created_at FROM scheduling_groups ORDER BY id")
            records = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "班组数据"
            
            headers = ["班组名称", "描述", "创建时间"]
            ws.append(headers)
            
            for record in records:
                row = [
                    record.get('name', ''),
                    record.get('description', ''),
                    record['created_at'].strftime('%Y-%m-%d %H:%M:%S') if record.get('created_at') else ''
                ]
                ws.append(row)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"班组数据_{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib').parse.quote(filename)}"}
            )
    except Exception as e:
        logger.error(f"导出班组数据失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/groups/import/")
async def import_scheduling_groups(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """导入班组数据"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            return {"code": 400, "message": "只支持Excel文件格式(.xlsx, .xls)"}
        
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        count = 0
        with conn.cursor() as cursor:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                name = str(row[0]).strip() if row[0] else ''
                description = str(row[1]).strip() if row[1] else ''
                
                if not name:
                    continue
                
                try:
                    cursor.execute(
                        "INSERT INTO scheduling_groups (name, description) VALUES (%s, %s)",
                        (name, description)
                    )
                    count += 1
                except Exception as e:
                    logger.warning(f"导入班组跳过: {name}, 错误: {e}")
                    continue
            
            conn.commit()
        
        return {"code": 200, "message": "导入成功", "data": {"count": count}}
    except Exception as e:
        logger.error(f"导入班组数据失败: {e}")
        return {"code": 500, "message": f"导入失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/shifts/")
async def get_scheduling_shifts(
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """获取班次列表"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": []}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM scheduling_shifts ORDER BY start_time")
            shifts = cursor.fetchall()
            for shift in shifts:
                if shift['start_time']:
                    shift['start_time'] = format_time_value(shift['start_time'])
                if shift['end_time']:
                    shift['end_time'] = format_time_value(shift['end_time'])
            return {"code": 200, "message": "success", "data": shifts}
    except Exception as e:
        logger.error(f"获取班次列表失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/shifts/")
async def create_scheduling_shift(
    shift: SchedulingShiftCreate,
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """创建班次"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO scheduling_shifts (name, start_time, end_time, description) VALUES (%s, %s, %s, %s)",
                (shift.name, shift.start_time, shift.end_time, shift.description)
            )
            conn.commit()
            return {"code": 200, "message": "创建成功", "data": {"id": cursor.lastrowid}}
    except Exception as e:
        logger.error(f"创建班次失败: {e}")
        return {"code": 500, "message": f"创建失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.put("/api/scheduling/shifts/{shift_id}")
async def update_scheduling_shift(
    shift_id: int,
    shift: SchedulingShiftUpdate,
    user: dict = Depends(require_permission('data-records:scheduling:edit'))
):
    """更新班次"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            update_fields = []
            params = []
            if shift.name is not None:
                update_fields.append("name = %s")
                params.append(shift.name)
            if shift.start_time is not None:
                update_fields.append("start_time = %s")
                params.append(shift.start_time)
            if shift.end_time is not None:
                update_fields.append("end_time = %s")
                params.append(shift.end_time)
            if shift.description is not None:
                update_fields.append("description = %s")
                params.append(shift.description)
            
            if not update_fields:
                return {"code": 400, "message": "没有需要更新的字段"}
            
            params.append(shift_id)
            sql = f"UPDATE scheduling_shifts SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(sql, params)
            conn.commit()
            return {"code": 200, "message": "更新成功"}
    except Exception as e:
        logger.error(f"更新班次失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/scheduling/shifts/{shift_id}")
async def delete_scheduling_shift(
    shift_id: int,
    user: dict = Depends(require_permission('data-records:scheduling:delete'))
):
    """删除班次"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM scheduling_shifts WHERE id = %s", (shift_id,))
            conn.commit()
            return {"code": 200, "message": "删除成功"}
    except Exception as e:
        logger.error(f"删除班次失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/shifts/export/")
async def export_scheduling_shifts(
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """导出班次数据"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到数据库")
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT name, start_time, end_time, description FROM scheduling_shifts ORDER BY start_time")
            records = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "班次数据"
            
            headers = ["班次名称", "开始时间", "结束时间", "描述"]
            ws.append(headers)
            
            for record in records:
                row = [
                    record.get('name', ''),
                    format_time_value(record.get('start_time')),
                    format_time_value(record.get('end_time')),
                    record.get('description', '')
                ]
                ws.append(row)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"班次数据_{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib').parse.quote(filename)}"}
            )
    except Exception as e:
        logger.error(f"导出班次数据失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/shifts/import/")
async def import_scheduling_shifts(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """导入班次数据"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            return {"code": 400, "message": "只支持Excel文件格式(.xlsx, .xls)"}
        
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        count = 0
        with conn.cursor() as cursor:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                name = str(row[0]).strip() if row[0] else ''
                start_time = str(row[1]).strip() if row[1] else ''
                end_time = str(row[2]).strip() if row[2] else ''
                description = str(row[3]).strip() if row[3] else ''
                
                if not name or not start_time or not end_time:
                    continue
                
                try:
                    cursor.execute(
                        "INSERT INTO scheduling_shifts (name, start_time, end_time, description) VALUES (%s, %s, %s, %s)",
                        (name, start_time, end_time, description)
                    )
                    count += 1
                except Exception as e:
                    logger.warning(f"导入班次跳过: {name}, 错误: {e}")
                    continue
            
            conn.commit()
        
        return {"code": 200, "message": "导入成功", "data": {"count": count}}
    except Exception as e:
        logger.error(f"导入班次数据失败: {e}")
        return {"code": 500, "message": f"导入失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/staff/")
async def get_scheduling_staff(
    group_id: int = None,
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """获取人员列表"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": []}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            if group_id:
                cursor.execute(
                    "SELECT s.*, g.name as group_name FROM scheduling_staff s LEFT JOIN scheduling_groups g ON s.group_id = g.id WHERE s.group_id = %s ORDER BY s.id",
                    (group_id,)
                )
            else:
                cursor.execute("SELECT s.*, g.name as group_name FROM scheduling_staff s LEFT JOIN scheduling_groups g ON s.group_id = g.id ORDER BY s.id")
            staff = cursor.fetchall()
            return {"code": 200, "message": "success", "data": staff}
    except Exception as e:
        logger.error(f"获取人员列表失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/staff/")
async def create_scheduling_staff(
    staff: SchedulingStaffCreate,
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """创建人员"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO scheduling_staff (name, group_id, phone) VALUES (%s, %s, %s)",
                (staff.name, staff.group_id, staff.phone)
            )
            conn.commit()
            return {"code": 200, "message": "创建成功", "data": {"id": cursor.lastrowid}}
    except Exception as e:
        logger.error(f"创建人员失败: {e}")
        return {"code": 500, "message": f"创建失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.put("/api/scheduling/staff/{staff_id}")
async def update_scheduling_staff(
    staff_id: int,
    staff: SchedulingStaffUpdate,
    user: dict = Depends(require_permission('data-records:scheduling:edit'))
):
    """更新人员"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            update_fields = []
            params = []
            if staff.name is not None:
                update_fields.append("name = %s")
                params.append(staff.name)
            if staff.group_id is not None:
                update_fields.append("group_id = %s")
                params.append(staff.group_id)
            if staff.phone is not None:
                update_fields.append("phone = %s")
                params.append(staff.phone)
            if staff.status is not None:
                update_fields.append("status = %s")
                params.append(staff.status)
            
            if not update_fields:
                return {"code": 400, "message": "没有需要更新的字段"}
            
            params.append(staff_id)
            sql = f"UPDATE scheduling_staff SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(sql, params)
            conn.commit()
            return {"code": 200, "message": "更新成功"}
    except Exception as e:
        logger.error(f"更新人员失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/scheduling/staff/{staff_id}")
async def delete_scheduling_staff(
    staff_id: int,
    user: dict = Depends(require_permission('data-records:scheduling:delete'))
):
    """删除人员"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM scheduling_staff WHERE id = %s", (staff_id,))
            conn.commit()
            return {"code": 200, "message": "删除成功"}
    except Exception as e:
        logger.error(f"删除人员失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/staff/export/")
async def export_scheduling_staff(
    group_id: int = None,
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """导出人员数据"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到数据库")
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            where_sql = "1=1"
            params = []
            if group_id:
                where_sql = "staff.group_id = %s"
                params.append(group_id)
            
            sql = f"""
                SELECT staff.name, g.name as group_name, staff.phone, 
                       CASE staff.status WHEN 1 THEN '在职' ELSE '离职' END as status_text
                FROM scheduling_staff staff
                LEFT JOIN scheduling_groups g ON staff.group_id = g.id
                WHERE {where_sql}
                ORDER BY staff.id
            """
            cursor.execute(sql, params)
            records = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "人员数据"
            
            headers = ["姓名", "所属班组", "联系电话", "状态"]
            ws.append(headers)
            
            for record in records:
                row = [
                    record.get('name', ''),
                    record.get('group_name', ''),
                    record.get('phone', ''),
                    record.get('status_text', '')
                ]
                ws.append(row)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"人员数据_{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib').parse.quote(filename)}"}
            )
    except Exception as e:
        logger.error(f"导出人员数据失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/staff/import/")
async def import_scheduling_staff(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """导入人员数据"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            return {"code": 400, "message": "只支持Excel文件格式(.xlsx, .xls)"}
        
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        group_map = {}
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT id, name FROM scheduling_groups")
            for row in cursor.fetchall():
                group_map[row['name']] = row['id']
        
        count = 0
        with conn.cursor() as cursor:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                name = str(row[0]).strip() if row[0] else ''
                group_name = str(row[1]).strip() if row[1] else ''
                phone = str(row[2]).strip() if row[2] else ''
                status_text = str(row[3]).strip() if row[3] else '在职'
                
                if not name:
                    continue
                
                group_id = group_map.get(group_name, None)
                status = 1 if status_text == '在职' else 0
                
                try:
                    cursor.execute(
                        "INSERT INTO scheduling_staff (name, group_id, phone, status) VALUES (%s, %s, %s, %s)",
                        (name, group_id, phone, status)
                    )
                    count += 1
                except Exception as e:
                    logger.warning(f"导入人员跳过: {name}, 错误: {e}")
                    continue
            
            conn.commit()
        
        return {"code": 200, "message": "导入成功", "data": {"count": count}}
    except Exception as e:
        logger.error(f"导入人员数据失败: {e}")
        return {"code": 500, "message": f"导入失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/records/")
async def get_scheduling_records(
    year: int = None,
    month: int = None,
    group_id: int = None,
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """获取排班记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库", "data": []}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            where_clauses = []
            params = []
            
            if year and month:
                where_clauses.append("YEAR(sr.schedule_date) = %s AND MONTH(sr.schedule_date) = %s")
                params.extend([year, month])
            if group_id:
                where_clauses.append("sr.group_id = %s")
                params.append(group_id)
            
            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            sql = f"""
                SELECT sr.*, 
                       s.name as staff_name, s.employee_id,
                       g.name as group_name,
                       sh.name as shift_name, sh.start_time, sh.end_time
                FROM scheduling_records sr
                LEFT JOIN scheduling_staff s ON sr.staff_id = s.id
                LEFT JOIN scheduling_groups g ON sr.group_id = g.id
                LEFT JOIN scheduling_shifts sh ON sr.shift_id = sh.id
                WHERE {where_sql}
                ORDER BY sr.schedule_date, sh.start_time
            """
            cursor.execute(sql, params)
            records = cursor.fetchall()
            
            for record in records:
                if record['schedule_date']:
                    record['schedule_date'] = record['schedule_date'].strftime('%Y-%m-%d')
                if record['start_time']:
                    record['start_time'] = format_time_value(record['start_time'])
                if record['end_time']:
                    record['end_time'] = format_time_value(record['end_time'])
            
            return {"code": 200, "message": "success", "data": records}
    except Exception as e:
        logger.error(f"获取排班记录失败: {e}")
        return {"code": 500, "message": f"获取失败: {str(e)}", "data": []}
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/records/")
async def create_scheduling_record(
    record: SchedulingRecordCreate,
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """创建排班记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            if record.staff_id:
                cursor.execute(
                    """INSERT INTO scheduling_records (staff_id, group_id, shift_id, schedule_date, remark) 
                       VALUES (%s, %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE remark = VALUES(remark)""",
                    (record.staff_id, record.group_id, record.shift_id, record.schedule_date, record.remark)
                )
                conn.commit()
                return {"code": 200, "message": "创建成功", "data": {"id": cursor.lastrowid}}
            else:
                cursor.execute(
                    "SELECT id FROM scheduling_staff WHERE group_id = %s AND status = 1",
                    (record.group_id,)
                )
                staff_list = cursor.fetchall()
                
                if not staff_list:
                    return {"code": 400, "message": "该班组没有在职人员"}
                
                inserted_count = 0
                for staff in staff_list:
                    try:
                        cursor.execute(
                            """INSERT INTO scheduling_records (staff_id, group_id, shift_id, schedule_date, remark) 
                               VALUES (%s, %s, %s, %s, %s)
                               ON DUPLICATE KEY UPDATE remark = VALUES(remark)""",
                            (staff[0], record.group_id, record.shift_id, record.schedule_date, record.remark)
                        )
                        inserted_count += 1
                    except Exception as e:
                        logger.warning(f"为人员 {staff[0]} 创建排班记录失败: {e}")
                        continue
                
                conn.commit()
                return {"code": 200, "message": f"成功为班组所有成员创建排班记录，共 {inserted_count} 条", "data": {"count": inserted_count}}
    except Exception as e:
        logger.error(f"创建排班记录失败: {e}")
        return {"code": 500, "message": f"创建失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.put("/api/scheduling/records/{record_id}")
async def update_scheduling_record(
    record_id: int,
    record: SchedulingRecordUpdate,
    user: dict = Depends(require_permission('data-records:scheduling:edit'))
):
    """更新排班记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            update_fields = []
            params = []
            if record.staff_id is not None:
                update_fields.append("staff_id = %s")
                params.append(record.staff_id)
            if record.group_id is not None:
                update_fields.append("group_id = %s")
                params.append(record.group_id)
            if record.shift_id is not None:
                update_fields.append("shift_id = %s")
                params.append(record.shift_id)
            if record.schedule_date is not None:
                update_fields.append("schedule_date = %s")
                params.append(record.schedule_date)
            if record.remark is not None:
                update_fields.append("remark = %s")
                params.append(record.remark)
            
            if not update_fields:
                return {"code": 400, "message": "没有需要更新的字段"}
            
            params.append(record_id)
            sql = f"UPDATE scheduling_records SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(sql, params)
            conn.commit()
            return {"code": 200, "message": "更新成功"}
    except Exception as e:
        logger.error(f"更新排班记录失败: {e}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/scheduling/records/{record_id}")
async def delete_scheduling_record(
    record_id: int,
    user: dict = Depends(require_permission('data-records:scheduling:delete'))
):
    """删除排班记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM scheduling_records WHERE id = %s", (record_id,))
            conn.commit()
            return {"code": 200, "message": "删除成功"}
    except Exception as e:
        logger.error(f"删除排班记录失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.delete("/api/scheduling/records/")
async def delete_scheduling_records_by_date(
    year: int,
    month: int,
    group_id: int = None,
    user: dict = Depends(require_permission('data-records:scheduling:delete'))
):
    """删除指定月份的排班记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor() as cursor:
            if group_id:
                cursor.execute(
                    "DELETE FROM scheduling_records WHERE YEAR(schedule_date) = %s AND MONTH(schedule_date) = %s AND group_id = %s",
                    (year, month, group_id)
                )
            else:
                cursor.execute(
                    "DELETE FROM scheduling_records WHERE YEAR(schedule_date) = %s AND MONTH(schedule_date) = %s",
                    (year, month)
                )
            conn.commit()
            return {"code": 200, "message": f"成功删除 {cursor.rowcount} 条记录"}
    except Exception as e:
        logger.error(f"删除排班记录失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/scheduling/export/")
async def export_scheduling_records(
    year: int,
    month: int,
    group_id: int = None,
    user: dict = Depends(require_permission('data-records:scheduling:view'))
):
    """导出排班记录"""
    try:
        conn = create_db_connection("USER_DB", config)
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接到数据库")
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            where_clauses = ["YEAR(sr.schedule_date) = %s", "MONTH(sr.schedule_date) = %s"]
            params = [year, month]
            
            if group_id:
                where_clauses.append("sr.group_id = %s")
                params.append(group_id)
            
            where_sql = " AND ".join(where_clauses)
            
            sql = f"""
                SELECT sr.schedule_date, s.name as staff_name, 
                       g.name as group_name, sh.name as shift_name, 
                       sh.start_time, sh.end_time, sr.remark
                FROM scheduling_records sr
                LEFT JOIN scheduling_staff s ON sr.staff_id = s.id
                LEFT JOIN scheduling_groups g ON sr.group_id = g.id
                LEFT JOIN scheduling_shifts sh ON sr.shift_id = sh.id
                WHERE {where_sql}
                ORDER BY sr.schedule_date, sh.start_time, s.name
            """
            cursor.execute(sql, params)
            records = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{year}年{month}月排班"
            
            headers = ["日期", "姓名", "班组", "班次", "开始时间", "结束时间", "备注"]
            ws.append(headers)
            
            for record in records:
                row = [
                    record['schedule_date'].strftime('%Y-%m-%d') if record.get('schedule_date') else '',
                    record.get('staff_name', ''),
                    record.get('group_name', ''),
                    record.get('shift_name', ''),
                    format_time_value(record.get('start_time')),
                    format_time_value(record.get('end_time')),
                    record.get('remark', '')
                ]
                ws.append(row)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"排班表_{year}年{month}月_{__import__('datetime').datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib').parse.quote(filename)}"}
            )
    except Exception as e:
        logger.error(f"导出排班记录失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/scheduling/import/")
async def import_scheduling_records(
    file: UploadFile = File(...),
    user: dict = Depends(require_permission('data-records:scheduling:create'))
):
    """导入排班记录"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            return {"code": 400, "message": "只支持Excel文件格式(.xlsx, .xls)"}
        
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return {"code": 400, "message": "Excel文件中没有数据"}
        
        conn = create_db_connection("USER_DB", config)
        if not conn:
            return {"code": 500, "message": "无法连接到数据库"}
        
        with conn.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT id, name FROM scheduling_groups")
            groups = {g['name']: g['id'] for g in cursor.fetchall()}
            
            cursor.execute("SELECT id, name FROM scheduling_shifts")
            shifts = {s['name']: s['id'] for s in cursor.fetchall()}
            
            cursor.execute("SELECT id, name FROM scheduling_staff")
            staff_list = cursor.fetchall()
            staff_by_name = {s['name']: s['id'] for s in staff_list}
        
        imported_count = 0
        with conn.cursor() as cursor:
            for row in rows:
                if not row or not any(row):
                    continue
                
                try:
                    date_val = row[0] if len(row) > 0 else None
                    if hasattr(date_val, 'strftime'):
                        date_val = date_val.strftime('%Y-%m-%d')
                    elif isinstance(date_val, str):
                        pass
                    else:
                        continue
                    
                    staff_name = str(row[1]) if len(row) > 1 and row[1] else None
                    group_name = str(row[2]) if len(row) > 2 and row[2] else None
                    shift_name = str(row[3]) if len(row) > 3 and row[3] else None
                    remark = str(row[6]) if len(row) > 6 and row[6] else None
                    
                    staff_id = staff_by_name.get(staff_name)
                    group_id = groups.get(group_name)
                    shift_id = shifts.get(shift_name)
                    
                    if not staff_id or not group_id or not shift_id:
                        continue
                    
                    cursor.execute(
                        """INSERT INTO scheduling_records (staff_id, group_id, shift_id, schedule_date, remark) 
                           VALUES (%s, %s, %s, %s, %s)
                           ON DUPLICATE KEY UPDATE remark = VALUES(remark)""",
                        (staff_id, group_id, shift_id, date_val, remark)
                    )
                    imported_count += 1
                except Exception as e:
                    logger.warning(f"导入行失败: {e}, 行数据: {row}")
                    continue
            
            conn.commit()
        
        return {"code": 200, "message": "导入成功", "data": {"count": imported_count}}
    except Exception as e:
        logger.error(f"导入排班记录失败: {e}")
        return {"code": 500, "message": f"导入失败: {str(e)}"}
    finally:
        if 'conn' in locals():
            conn.close()

GUI_SERVICE_URL = "http://172.32.48.239:9000"

class CloudPortalLoginRequest(BaseModel):
    session_id: str
    username: str
    password: str
    captcha: str
    uuid: str

class CloudPortalQueryRequest(BaseModel):
    session_id: str
    query_params: dict

class CloudPortalLogoutRequest(BaseModel):
    session_id: str

@app.get("/api/cloud-portal/captcha")
async def get_cloud_portal_captcha(session_id: Optional[str] = None):
    try:
        params = {}
        if session_id:
            params['session_id'] = session_id
        
        logger.info(f"[云门户验证码] 开始请求 - session_id: {session_id}, 目标: {GUI_SERVICE_URL}/api/portal/captcha")
        
        loop = asyncio.get_event_loop()
        response = await asyncio.to_thread(
            requests.get,
            f"{GUI_SERVICE_URL}/api/portal/captcha",
            params=params,
            timeout=30
        )
        
        logger.info(f"[云门户验证码] 请求成功 - 状态码: {response.status_code}")
        return response.json()
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[云门户验证码] 连接失败 - 错误: {e}")
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务，请确保服务已启动"}
        )
    except requests.exceptions.Timeout:
        logger.error("[云门户验证码] 请求超时")
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"[云门户验证码] 异常 - 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"获取验证码失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/login")
async def cloud_portal_login(request: CloudPortalLoginRequest):
    try:
        logger.info(f"[云门户登录] 开始请求 - session_id: {request.session_id}")

        response = await asyncio.to_thread(
            requests.post,
            f"{GUI_SERVICE_URL}/api/portal/login",
            json={
                'session_id': request.session_id,
                'username': request.username,
                'password': request.password,
                'captcha': request.captcha,
                'uuid': request.uuid
            },
            timeout=20
        )

        logger.info(f"[云门户登录] 请求成功 - 状态码: {response.status_code}")
        return response.json()
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[云门户登录] 连接失败 - 错误: {e}")
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务，请确保服务已启动"}
        )
    except requests.exceptions.Timeout:
        logger.error("[云门户登录] 请求超时")
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"[云门户登录] 异常 - 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"登录失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/query")
async def cloud_portal_query(request: CloudPortalQueryRequest):
    try:
        logger.info(f"[云门户查询] 开始请求 - session_id: {request.session_id}")

        response = await asyncio.to_thread(
            requests.post,
            f"{GUI_SERVICE_URL}/api/portal/query",
            json={
                'session_id': request.session_id,
                'query_params': request.query_params
            },
            timeout=45
        )

        logger.info(f"[云门户查询] 请求成功 - 状态码: {response.status_code}")
        return response.json()
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[云门户查询] 连接失败 - 错误: {e}")
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务，请确保服务已启动"}
        )
    except requests.exceptions.Timeout:
        logger.error("[云门户查询] 请求超时")
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"[云门户查询] 异常 - 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

@app.get("/api/cloud-portal/status")
async def cloud_portal_status(session_id: str):
    try:
        logger.info(f"[云门户状态] 开始请求 - session_id: {session_id}")

        response = await asyncio.to_thread(
            requests.get,
            f"{GUI_SERVICE_URL}/api/portal/status",
            params={'session_id': session_id},
            timeout=15
        )

        logger.info(f"[云门户状态] 请求成功 - 状态码: {response.status_code}")
        return response.json()
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[云门户状态] 连接失败 - 错误: {e}")
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务，请确保服务已启动"}
        )
    except requests.exceptions.Timeout:
        logger.error("[云门户状态] 请求超时")
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"[云门户状态] 异常 - 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"获取状态失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/logout")
async def cloud_portal_logout(request: CloudPortalLogoutRequest):
    try:
        logger.info(f"[云门户登出] 开始请求 - session_id: {request.session_id}")

        response = await asyncio.to_thread(
            requests.post,
            f"{GUI_SERVICE_URL}/api/portal/logout",
            json={'session_id': request.session_id},
            timeout=15
        )

        logger.info(f"[云门户登出] 请求成功 - 状态码: {response.status_code}")
        return response.json()
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[云门户登出] 连接失败 - 错误: {e}")
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务，请确保服务已启动"}
        )
    except requests.exceptions.Timeout:
        logger.error("[云门户登出] 请求超时")
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"[云门户登出] 异常 - 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"登出失败: {str(e)}"}
        )

@app.get("/api/cloud-portal/health")
async def cloud_portal_health():
    try:
        logger.info("[云门户健康检查] 开始请求")

        response = await asyncio.to_thread(
            requests.get,
            f"{GUI_SERVICE_URL}/api/portal/health",
            timeout=10
        )

        logger.info(f"[云门户健康检查] 请求成功 - 状态码: {response.status_code}")
        return response.json()
    except requests.exceptions.ConnectionError as e:
        logger.error(f"[云门户健康检查] 连接失败 - 错误: {e}")
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "云门户查询服务未启动", "data": {"status": "offline"}}
        )
    except requests.exceptions.Timeout:
        logger.error("[云门户健康检查] 请求超时")
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "云门户查询服务响应超时", "data": {"status": "timeout"}}
        )
    except Exception as e:
        logger.error(f"[云门户健康检查] 异常 - 错误: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"健康检查失败: {str(e)}", "data": {"status": "error"}}
        )

class AIAuditBatchQueryRequest(BaseModel):
    session_id: str
    plate_number: str
    entry_time: str
    gate_time: str
    pass_id: Optional[str] = None
    hours: Optional[int] = 5
    rows: Optional[int] = 40

class AIAuditQueryRequest(BaseModel):
    session_id: str
    plate_number: Optional[str] = None
    query_value: Optional[str] = None
    vehicle_or_pass_id: Optional[str] = None
    start_time: str
    end_time: str
    trade_type: Optional[int] = 1

class AIAuditSelectImagesRequest(BaseModel):
    images: List[dict]
    gantry_ids: List[str]

class SaveImagesRequest(BaseModel):
    table_name: str
    record_id: str
    image1_base64: Optional[str] = None
    image2_base64: Optional[str] = None
    review_status: Optional[str] = None
    check_pass_id: Optional[str] = None
    special_situation: Optional[str] = None
    check_split: Optional[str] = None
    remark: Optional[str] = None
    clear_empty: bool = False

@app.post("/api/cloud-portal/ai-audit/vehicle-images")
async def ai_audit_vehicle_images(request: AIAuditQueryRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/vehicle-images",
            json={
                'session_id': request.session_id,
                'plate_number': request.plate_number,
                'start_time': request.start_time,
                'end_time': request.end_time
            },
            timeout=60
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"车辆图库查询失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/gantry-trade")
async def ai_audit_gantry_trade(request: AIAuditQueryRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/gantry-trade",
            json={
                'session_id': request.session_id,
                'query_value': request.query_value,
                'start_time': request.start_time,
                'end_time': request.end_time
            },
            timeout=60
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"门架交易查询失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/gantry-plate")
async def ai_audit_gantry_plate(request: AIAuditQueryRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/gantry-plate",
            json={
                'session_id': request.session_id,
                'plate_number': request.plate_number,
                'start_time': request.start_time,
                'end_time': request.end_time
            },
            timeout=60
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"门架牌识查询失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/exit-trade")
async def ai_audit_exit_trade(request: AIAuditQueryRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/exit-trade",
            json={
                'session_id': request.session_id,
                'query_value': request.query_value,
                'start_time': request.start_time,
                'end_time': request.end_time,
                'trade_type': request.trade_type
            },
            timeout=60
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"出口交易查询失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/suspected-car")
async def ai_audit_suspected_car(request: AIAuditQueryRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/suspected-car",
            json={
                'session_id': request.session_id,
                'vehicle_or_pass_id': request.vehicle_or_pass_id,
                'start_time': request.start_time,
                'end_time': request.end_time
            },
            timeout=60
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"疑难车牌追查失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/batch-query")
async def ai_audit_batch_query(request: AIAuditBatchQueryRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/batch-query",
            json={
                'session_id': request.session_id,
                'plate_number': request.plate_number,
                'entry_time': request.entry_time,
                'gate_time': request.gate_time,
                'pass_id': request.pass_id,
                'hours': request.hours,
                'rows': request.rows
            },
            timeout=120
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"批量查询失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"查询失败: {str(e)}"}
        )

class OriginalImageRequest(BaseModel):
    session_id: str
    picture_path: str

@app.post("/api/cloud-portal/ai-audit/original-image")
async def ai_audit_original_image(request: OriginalImageRequest):
    try:
        loop = asyncio.get_event_loop()
        response = await asyncio.to_thread(
            requests.post,
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/original-image",
            json={
                'session_id': request.session_id,
                'picture_path': request.picture_path
            },
            timeout=30
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"获取高清原图失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"获取原图失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/select-images")
async def ai_audit_select_images(request: AIAuditSelectImagesRequest):
    try:
        response = requests.post(
            f"{GUI_SERVICE_URL}/api/portal/ai-audit/select-images",
            json={
                'images': request.images,
                'gantry_ids': request.gantry_ids
            },
            timeout=30
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"图片选取失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"图片选取失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/save-images")
async def ai_audit_save_images(request: SaveImagesRequest):
    try:
        conn = create_db_connection('CHECK_DATA_DB', config)
        cursor = conn.cursor()
        
        update_fields = []
        params = []
        
        field_mapping = {
            'image1_base64': '查核资料1',
            'image2_base64': '查核资料2',
            'review_status': '复核情况',
            'check_pass_id': '核查通行标识',
            'special_situation': '特情',
            'check_split': '核查拆分',
            'remark': '备注'
        }
        
        for field_name, db_column in field_mapping.items():
            value = getattr(request, field_name, None)
            if request.clear_empty or value:
                update_fields.append(f"`{db_column}` = %s")
                params.append(value if value else None)
        
        if not update_fields:
            return {"code": 400, "message": "没有需要保存的数据"}
        
        params.append(request.record_id)
        
        sql = f"UPDATE `{request.table_name}` SET {', '.join(update_fields)} WHERE `通行标识ID` = %s"
        cursor.execute(sql, params)
        conn.commit()
        
        affected_rows = cursor.rowcount
        cursor.close()
        conn.close()
        
        return {
            "code": 200,
            "message": f"成功保存 {affected_rows} 条记录",
            "data": {"affected_rows": affected_rows}
        }
    except Exception as e:
        logger.error(f"保存图片失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"保存图片失败: {str(e)}"}
        )

class BranchCentersRequest(BaseModel):
    session_id: str = None

class RoadSectionsRequest(BaseModel):
    session_id: str = None
    center_no: str

class GantryListRequest(BaseModel):
    session_id: str = None
    road_section_no: str

@app.post("/api/cloud-portal/ai-audit/branch-centers")
async def ai_audit_branch_centers(request: BranchCentersRequest):
    try:
        gui_service_url = "http://127.0.0.1:5000/api/portal/ai-audit/branch-centers"
        response = requests.post(
            gui_service_url,
            json={"session_id": request.session_id},
            timeout=30
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"获取分中心列表失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"获取分中心列表失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/road-sections")
async def ai_audit_road_sections(request: RoadSectionsRequest):
    try:
        gui_service_url = "http://127.0.0.1:5000/api/portal/ai-audit/road-sections"
        response = requests.post(
            gui_service_url,
            json={"session_id": request.session_id, "center_no": request.center_no},
            timeout=30
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"获取路段列表失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"获取路段列表失败: {str(e)}"}
        )

@app.post("/api/cloud-portal/ai-audit/gantry-list")
async def ai_audit_gantry_list(request: GantryListRequest):
    try:
        gui_service_url = "http://127.0.0.1:5000/api/portal/ai-audit/gantry-list"
        response = requests.post(
            gui_service_url,
            json={"session_id": request.session_id, "road_section_no": request.road_section_no},
            timeout=30
        )
        return response.json()
    except requests.exceptions.ConnectionError:
        return JSONResponse(
            status_code=503,
            content={"code": 503, "message": "无法连接到云门户查询服务"}
        )
    except requests.exceptions.Timeout:
        return JSONResponse(
            status_code=504,
            content={"code": 504, "message": "连接云门户查询服务超时"}
        )
    except Exception as e:
        logger.error(f"获取门架列表失败: {e}")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"获取门架列表失败: {str(e)}"}
        )

@app.on_event("shutdown")
async def shutdown_event():
    """应用关闭时清理连接池"""
    logger.info("应用关闭，正在清理数据库连接池...")
    try:
        close_all_pools()
        logger.info("数据库连接池已清理完成")
    except Exception as e:
        logger.error(f"清理连接池时出错: {e}", exc_info=True)

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000);