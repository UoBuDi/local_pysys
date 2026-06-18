"""
追查详单路由
包含加入追查、查询列表、提交复核、删除记录、导出数据等接口
"""

import logging
import io

import pymysql
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime
from pydantic import BaseModel, Field

from core.dependencies import get_current_user, require_permission
from core.config import get_config as get_app_config

logger = logging.getLogger(__name__)

router = APIRouter()


class AddInvestigationRequest(BaseModel):
    """加入追查请求"""
    pass_id: str = Field(..., description="通行标识ID")
    plate_number: str = Field(..., description="车牌号码")


class ReviewInvestigationRequest(BaseModel):
    """提交复核请求"""
    id: int = Field(..., description="追查详单记录ID")
    review_result: str = Field(..., max_length=200, description="复核结果（最多200字符）")


def _get_db_conn():
    """获取system_db数据库连接"""
    config = get_app_config()
    from database import get_db_connection
    return get_db_connection("USER_DB", config)


@router.post("/api/investigation/add")
async def add_to_investigation(
    request: AddInvestigationRequest,
    user: dict = Depends(get_current_user)
):
    """将通行标识ID和车牌号码加入追查详单"""
    try:
        username = user.get("username", "unknown")

        with _get_db_conn() as conn:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 检查是否已存在相同记录（未复核的重复记录）
                cursor.execute(
                    "SELECT id FROM investigation_details WHERE pass_id = %s AND plate_number = %s AND review_result IS NULL",
                    (request.pass_id, request.plate_number)
                )
                if cursor.fetchone():
                    return {"code": 400, "message": "该通行标识ID和车牌号码已存在未复核的追查记录"}

                cursor.execute(
                    """INSERT INTO investigation_details (pass_id, plate_number, add_time, created_by)
                       VALUES (%s, %s, NOW(), %s)""",
                    (request.pass_id, request.plate_number, username)
                )
                conn.commit()

                # 返回新插入的记录
                cursor.execute(
                    "SELECT * FROM investigation_details WHERE id = LAST_INSERT_ID()"
                )
                new_record = cursor.fetchone()

                # 格式化时间字段
                if new_record:
                    for key in ['add_time', 'review_time', 'created_at', 'updated_at']:
                        if new_record.get(key) and isinstance(new_record[key], datetime):
                            new_record[key] = new_record[key].strftime('%Y-%m-%d %H:%M:%S')

                return {"code": 200, "message": "加入追查成功", "data": new_record}

    except Exception as e:
        logger.error(f"加入追查失败: {e}")
        return {"code": 500, "message": f"加入追查失败: {str(e)}"}


@router.get("/api/investigation/list")
async def get_investigation_list(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    pass_id: Optional[str] = Query(None, description="通行标识ID（模糊查询）"),
    plate_number: Optional[str] = Query(None, description="车牌号码（模糊查询）"),
    created_by: Optional[str] = Query(None, description="创建人"),
    review_status: Optional[str] = Query(None, description="复核状态：reviewed/unreviewed"),
    start_time: Optional[str] = Query(None, description="加入时间起（格式：2026-06-16 00:00:00）"),
    end_time: Optional[str] = Query(None, description="加入时间止（格式：2026-06-16 23:59:59）"),
    user: dict = Depends(get_current_user)
):
    """分页查询追查详单列表"""
    try:
        with _get_db_conn() as conn:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 构建查询条件
                conditions = []
                params = []

                if pass_id:
                    conditions.append("pass_id LIKE %s")
                    params.append(f"%{pass_id}%")
                if plate_number:
                    conditions.append("plate_number LIKE %s")
                    params.append(f"%{plate_number}%")
                if created_by:
                    conditions.append("created_by = %s")
                    params.append(created_by)
                if review_status == "reviewed":
                    conditions.append("review_result IS NOT NULL")
                elif review_status == "unreviewed":
                    conditions.append("review_result IS NULL")
                if start_time:
                    conditions.append("add_time >= %s")
                    params.append(start_time)
                if end_time:
                    conditions.append("add_time <= %s")
                    params.append(end_time)

                where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""

                # 查询总数
                cursor.execute(f"SELECT COUNT(*) as total FROM investigation_details{where_clause}", params)
                total = cursor.fetchone()['total']

                # 分页查询
                offset = (page - 1) * page_size
                cursor.execute(
                    f"SELECT * FROM investigation_details{where_clause} ORDER BY add_time DESC LIMIT %s OFFSET %s",
                    params + [page_size, offset]
                )
                records = cursor.fetchall()

                # 格式化时间字段
                for record in records:
                    for key in ['add_time', 'review_time', 'created_at', 'updated_at']:
                        if record.get(key) and isinstance(record[key], datetime):
                            record[key] = record[key].strftime('%Y-%m-%d %H:%M:%S')

                return {
                    "code": 200,
                    "message": "success",
                    "data": {
                        "records": records,
                        "total": total,
                        "page": page,
                        "page_size": page_size
                    }
                }

    except Exception as e:
        logger.error(f"查询追查详单列表失败: {e}")
        return {"code": 500, "message": f"查询失败: {str(e)}"}


@router.put("/api/investigation/review")
async def review_investigation(
    request: ReviewInvestigationRequest,
    user: dict = Depends(get_current_user)
):
    """提交复核结果"""
    try:
        username = user.get("username", "unknown")

        with _get_db_conn() as conn:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 检查记录是否存在
                cursor.execute("SELECT id FROM investigation_details WHERE id = %s", (request.id,))
                if not cursor.fetchone():
                    return {"code": 404, "message": "追查详单记录不存在"}

                cursor.execute(
                    """UPDATE investigation_details
                       SET review_result = %s, reviewed_by = %s, review_time = NOW()
                       WHERE id = %s""",
                    (request.review_result, username, request.id)
                )
                conn.commit()

                # 返回更新后的记录
                cursor.execute("SELECT * FROM investigation_details WHERE id = %s", (request.id,))
                updated_record = cursor.fetchone()

                if updated_record:
                    for key in ['add_time', 'review_time', 'created_at', 'updated_at']:
                        if updated_record.get(key) and isinstance(updated_record[key], datetime):
                            updated_record[key] = updated_record[key].strftime('%Y-%m-%d %H:%M:%S')

                return {"code": 200, "message": "复核成功", "data": updated_record}

    except Exception as e:
        logger.error(f"提交复核失败: {e}")
        return {"code": 500, "message": f"提交复核失败: {str(e)}"}


@router.delete("/api/investigation/{record_id}")
async def delete_investigation(
    record_id: int,
    user: dict = Depends(get_current_user)
):
    """删除追查详单记录"""
    try:
        with _get_db_conn() as conn:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 检查记录是否存在
                cursor.execute("SELECT id FROM investigation_details WHERE id = %s", (record_id,))
                if not cursor.fetchone():
                    return {"code": 404, "message": "追查详单记录不存在"}

                cursor.execute("DELETE FROM investigation_details WHERE id = %s", (record_id,))
                conn.commit()

                return {"code": 200, "message": "删除成功"}

    except Exception as e:
        logger.error(f"删除追查详单记录失败: {e}")
        return {"code": 500, "message": f"删除失败: {str(e)}"}


def _extract_yc_month_from_pass_id(pass_id: str) -> str:
    """从通行标识ID中提取年月，用于定位yc表

    通行标识ID格式中包含日期信息（YYYYMMDD），位于ID的固定位置。
    例如: 014301220423380731348020250103084015 中 20250103 → 2025-01

    Returns:
        str: 年月字符串，格式 YYYY-MM（如 '2025-01'）；无法提取时返回空字符串
    """
    import re
    # 通行标识ID中包含8位连续数字日期（YYYYMMDD），用正则提取
    match = re.search(r'(20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])', pass_id)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return ''


def _query_table_by_pass_id(cursor, table_name: str, pass_id: str) -> dict:
    """在指定表中按通行标识ID查询记录，返回 {columns, rows} 或 None"""
    # 检查表是否有通行标识ID列
    cursor.execute(f"DESCRIBE `{table_name}`")
    cols_info = cursor.fetchall()
    col_names = {row[0] for row in cols_info}
    if '通行标识ID' not in col_names:
        return None

    sql = f"SELECT * FROM `{table_name}` WHERE `通行标识ID` = %s LIMIT 50"
    cursor.execute(sql, (pass_id,))

    if not cursor.description:
        return None

    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    if not rows:
        return None

    return {"columns": columns, "rows": rows}


def _format_records(columns: list, rows: tuple) -> list:
    """将查询结果转换为字典列表，处理bytes和时间类型"""
    from core.utils import format_time_value
    time_columns = {
        '入口时间', '出口时间', '交易时间', '门架通行时间',
        '计费交易起点时间', '计费交易终点时间'
    }
    records = []
    for row in rows:
        row_dict = {}
        for i, col in enumerate(columns):
            val = row[i]
            if isinstance(val, bytes):
                try:
                    val = val.decode('utf-8')
                except UnicodeDecodeError:
                    val = val.hex()
            if col in time_columns:
                val = format_time_value(val)
            row_dict[col] = val
        records.append(row_dict)
    return records


@router.get("/api/investigation/pass-records")
async def get_pass_records(
    pass_id: str = Query(..., description="通行标识ID"),
    user: dict = Depends(get_current_user)
):
    """根据通行标识ID查询通行记录

    查询策略：
    1. 优先查询详单查询表（[DETAIL_QUERY]配置的表）
    2. 若无结果，从pass_id提取年月，优先查询对应月份的yc表
    3. 若仍无结果，遍历所有yc表查询
    """
    try:
        config = get_app_config()
        from database import get_db_connection

        with get_db_connection("CHECK_DATA_DB", config) as conn:
            with conn.cursor() as cursor:
                # ---- 第1步：查询详单查询表 ----
                detail_table = config.get('DETAIL_QUERY', 'table_name', fallback='')
                result_table = None

                if detail_table:
                    cursor.execute("SHOW TABLES LIKE %s", (detail_table,))
                    if cursor.fetchone():
                        query_result = _query_table_by_pass_id(cursor, detail_table, pass_id)
                        if query_result:
                            result_table = detail_table
                            columns = query_result["columns"]
                            rows = query_result["rows"]

                # ---- 第2步：详单查询表无结果，查询yc表 ----
                if not result_table:
                    # 获取check_data库中所有yc表（排除备份表）
                    cursor.execute("SHOW TABLES LIKE '%_yc'")
                    all_yc_tables = [row[0] for row in cursor.fetchall()
                                     if 'copy' not in row[0].lower()]

                    # 按表名倒序排列（优先查最新月份）
                    all_yc_tables.sort(reverse=True)

                    # 从pass_id提取年月，优先查询对应月份的yc表
                    target_month = _extract_yc_month_from_pass_id(pass_id)
                    priority_table = f"{target_month}_yc" if target_month else None

                    # 构建查询顺序：优先表在前
                    ordered_tables = []
                    if priority_table and priority_table in all_yc_tables:
                        ordered_tables.append(priority_table)
                    ordered_tables.extend(t for t in all_yc_tables if t != priority_table)

                    # 遍历查询
                    for yc_table in ordered_tables:
                        try:
                            query_result = _query_table_by_pass_id(cursor, yc_table, pass_id)
                            if query_result:
                                result_table = yc_table
                                columns = query_result["columns"]
                                rows = query_result["rows"]
                                break
                        except Exception as e:
                            logger.warning(f"查询yc表 {yc_table} 失败: {e}")
                            continue

                # ---- 无任何结果 ----
                if not result_table:
                    return {
                        "code": 200,
                        "message": "success",
                        "data": {
                            "records": [],
                            "columns": [],
                            "total": 0,
                            "table_name": ""
                        }
                    }

                # ---- 格式化并返回结果 ----
                records = _format_records(columns, rows)

                return {
                    "code": 200,
                    "message": "success",
                    "data": {
                        "records": records,
                        "columns": columns,
                        "total": len(records),
                        "table_name": result_table
                    }
                }

    except Exception as e:
        logger.error(f"查询通行记录失败: {e}")
        return {"code": 500, "message": f"查询通行记录失败: {str(e)}"}


@router.get("/api/investigation/export")
async def export_investigation(
    pass_id: Optional[str] = Query(None, description="通行标识ID（模糊查询）"),
    plate_number: Optional[str] = Query(None, description="车牌号码（模糊查询）"),
    created_by: Optional[str] = Query(None, description="创建人"),
    review_status: Optional[str] = Query(None, description="复核状态：reviewed/unreviewed"),
    start_time: Optional[str] = Query(None, description="加入时间起"),
    end_time: Optional[str] = Query(None, description="加入时间止"),
    user: dict = Depends(get_current_user)
):
    """导出追查详单数据为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return {"code": 500, "message": "缺少openpyxl库，无法导出Excel"}

    try:
        with _get_db_conn() as conn:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # 构建查询条件（与列表查询一致）
                conditions = []
                params = []

                if pass_id:
                    conditions.append("pass_id LIKE %s")
                    params.append(f"%{pass_id}%")
                if plate_number:
                    conditions.append("plate_number LIKE %s")
                    params.append(f"%{plate_number}%")
                if created_by:
                    conditions.append("created_by = %s")
                    params.append(created_by)
                if review_status == "reviewed":
                    conditions.append("review_result IS NOT NULL")
                elif review_status == "unreviewed":
                    conditions.append("review_result IS NULL")
                if start_time:
                    conditions.append("add_time >= %s")
                    params.append(start_time)
                if end_time:
                    conditions.append("add_time <= %s")
                    params.append(end_time)

                where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""

                cursor.execute(f"SELECT * FROM investigation_details{where_clause} ORDER BY add_time DESC", params)
                records = cursor.fetchall()

                # 格式化时间字段
                for record in records:
                    for key in ['add_time', 'review_time', 'created_at', 'updated_at']:
                        if record.get(key) and isinstance(record[key], datetime):
                            record[key] = record[key].strftime('%Y-%m-%d %H:%M:%S')

                # 创建Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "追查详单"

                # 表头
                headers = ["序号", "通行标识ID", "车牌号码", "加入时间", "创建人", "复核结果", "复核人", "复核时间"]
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # 数据行
                for idx, record in enumerate(records, 1):
                    row_data = [
                        idx,
                        record.get('pass_id', ''),
                        record.get('plate_number', ''),
                        record.get('add_time', ''),
                        record.get('created_by', ''),
                        record.get('review_result', ''),
                        record.get('reviewed_by', ''),
                        record.get('review_time', '')
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=idx + 1, column=col, value=value)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

                # 调整列宽
                column_widths = [8, 35, 15, 22, 12, 30, 12, 22]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

                # 写入内存
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                filename = f"追查详单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )

    except Exception as e:
        logger.error(f"导出追查详单失败: {e}")
        return {"code": 500, "message": f"导出失败: {str(e)}"}
